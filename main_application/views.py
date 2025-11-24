"""
MediSphere Hospital Management System - Authentication and Business logic  Views
File: main_application/views.py
"""

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from functools import wraps
from main_application.models import User, AuditLog
from django.utils import timezone



# =============================================================================
# CUSTOM DECORATORS FOR ROLE-BASED ACCESS CONTROL
# =============================================================================

def role_required(*allowed_roles):
    """
    Decorator to restrict access based on user roles.
    
    Usage:
        @role_required('DOCTOR', 'CLINICAL_OFFICER')
        def doctor_view(request):
            ...
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                messages.error(request, 'Please login to access this page.')
                return redirect('login')
            
            # Superusers can access everything
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            
            # Check if user has required role
            if request.user.role and request.user.role.name in allowed_roles:
                return view_func(request, *args, **kwargs)
            
            # Access denied
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('dashboard')
        
        return wrapper
    return decorator


def department_required(*allowed_departments):
    """
    Decorator to restrict access based on user departments.
    
    Usage:
        @department_required('LABORATORY', 'RADIOLOGY')
        def lab_view(request):
            ...
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                messages.error(request, 'Please login to access this page.')
                return redirect('login')
            
            # Superusers can access everything
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            
            # Check if user has required department
            if request.user.department and request.user.department.name in allowed_departments:
                return view_func(request, *args, **kwargs)
            
            # Access denied
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('dashboard')
        
        return wrapper
    return decorator


def permission_required(permission_field):
    """
    Decorator to check specific role permissions.
    
    Usage:
        @permission_required('can_prescribe')
        def prescription_view(request):
            ...
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                messages.error(request, 'Please login to access this page.')
                return redirect('login')
            
            # Superusers can access everything
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            
            # Check if user's role has the required permission
            if request.user.role and getattr(request.user.role, permission_field, False):
                return view_func(request, *args, **kwargs)
            
            # Access denied
            messages.error(request, f'You do not have {permission_field.replace("_", " ")} permission.')
            return redirect('dashboard')
        
        return wrapper
    return decorator


# =============================================================================
# AUTHENTICATION VIEWS
# =============================================================================

def login_view(request):
    """
    Handle user login and redirect to appropriate dashboard based on role.
    """
    # Redirect if already logged in
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')
        
        # Authenticate user
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_active:
                login(request, user)
                
                # Set session expiry
                if not remember_me:
                    request.session.set_expiry(0)  # Session expires when browser closes
                else:
                    request.session.set_expiry(1209600)  # 2 weeks
                
                # Log the login
                AuditLog.objects.create(
                    user=user,
                    action_type='LOGIN',
                    model_name='User',
                    object_id=user.id,
                    object_repr=str(user),
                    ip_address=get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:300]
                )
                
                # Success message
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                
                # Redirect to appropriate dashboard or next page
                next_url = request.GET.get('next')
                if next_url:
                    return redirect(next_url)
                
                return redirect('dashboard')
            else:
                messages.error(request, 'Your account has been deactivated. Please contact administration.')
        else:
            messages.error(request, 'Invalid username or password. Please try again.')
    
    context = {
        'page_title': 'Login - MediSphere Hospital'
    }
    return render(request, 'auth/login.html', context)


@login_required
def logout_view(request):
    """
    Handle user logout.
    """
    # Log the logout
    AuditLog.objects.create(
        user=request.user,
        action_type='LOGOUT',
        model_name='User',
        object_id=request.user.id,
        object_repr=str(request.user),
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:300]
    )
    
    username = request.user.get_full_name() or request.user.username
    logout(request)
    messages.success(request, f'You have been logged out successfully. See you soon, {username}!')
    return redirect('login')

"""
MediSphere Hospital Management System - Password Reset Views
File: main_application/views/password_views.py
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.tokens import default_token_generator
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.conf import settings
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from main_application.models import User, AuditLog, SMSLog
import random
import string

User = get_user_model()


# =============================================================================
# PASSWORD RESET REQUEST VIEW
# =============================================================================

def forgot_password_view(request):
    """
    Password reset request page.
    User enters email or username to receive reset link.
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        identifier = request.POST.get('identifier', '').strip()  # Email or username
        
        if not identifier:
            messages.error(request, 'Please enter your email or username.')
            return render(request, 'auth/forgot_password.html')
        
        # Try to find user by email or username
        user = None
        try:
            if '@' in identifier:
                # Looks like email
                user = User.objects.get(email__iexact=identifier, is_active=True)
            else:
                # Looks like username
                user = User.objects.get(username__iexact=identifier, is_active=True)
        except User.DoesNotExist:
            # Don't reveal whether user exists or not (security best practice)
            messages.success(
                request, 
                'If an account exists with that information, a password reset link has been sent.'
            )
            return redirect('forgot_password')
        
        if user:
            # Generate password reset token
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            # Build reset URL
            reset_url = request.build_absolute_uri(
                reverse('password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
            )
            
            # Send email
            if user.email:
                try:
                    subject = 'Password Reset - MediSphere Hospital'
                    message = render_to_string('auth/password_reset_email.html', {
                        'user': user,
                        'reset_url': reset_url,
                        'hospital_name': 'MediSphere Hospital'
                    })
                    
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [user.email],
                        fail_silently=False,
                        html_message=message
                    )
                    
                    # Log the action
                    AuditLog.objects.create(
                        user=user,
                        action_type='VIEW',
                        model_name='User',
                        object_id=user.id,
                        object_repr=f'Password reset requested for {user.username}',
                        ip_address=get_client_ip(request),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:300]
                    )
                    
                except Exception as e:
                    messages.error(request, 'Error sending email. Please contact IT support.')
                    return render(request, 'auth/forgot_password.html')
            
            # Send SMS if phone number exists
            if user.phone_number:
                send_password_reset_sms(user, reset_url)
            
            messages.success(
                request,
                'Password reset instructions have been sent to your email and phone.'
            )
            return redirect('login')
    
    context = {
        'page_title': 'Forgot Password - MediSphere Hospital'
    }
    return render(request, 'auth/forgot_password.html', context)


# =============================================================================
# PASSWORD RESET CONFIRM VIEW
# =============================================================================

def password_reset_confirm_view(request, uidb64, token):
    """
    Verify the password reset token and allow user to set new password.
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    try:
        # Decode the user ID
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    # Verify the token
    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            password1 = request.POST.get('password1')
            password2 = request.POST.get('password2')
            
            # Validate passwords
            if not password1 or not password2:
                messages.error(request, 'Please enter both password fields.')
                return render(request, 'auth/password_reset_confirm.html', {
                    'validlink': True,
                    'page_title': 'Reset Password'
                })
            
            if password1 != password2:
                messages.error(request, 'Passwords do not match.')
                return render(request, 'auth/password_reset_confirm.html', {
                    'validlink': True,
                    'page_title': 'Reset Password'
                })
            
            if len(password1) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
                return render(request, 'auth/password_reset_confirm.html', {
                    'validlink': True,
                    'page_title': 'Reset Password'
                })
            
            # Set the new password
            user.set_password(password1)
            user.save()
            
            # Log the password change
            AuditLog.objects.create(
                user=user,
                action_type='UPDATE',
                model_name='User',
                object_id=user.id,
                object_repr=f'Password reset completed for {user.username}',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:300]
            )
            
            messages.success(
                request,
                'Your password has been reset successfully! You can now login with your new password.'
            )
            return redirect('login')
        
        context = {
            'validlink': True,
            'user': user,
            'page_title': 'Reset Password - MediSphere Hospital'
        }
        return render(request, 'auth/password_reset_confirm.html', context)
    else:
        # Invalid or expired token
        messages.error(
            request,
            'This password reset link is invalid or has expired. Please request a new one.'
        )
        return redirect('forgot_password')


# =============================================================================
# CHANGE PASSWORD VIEW (For Logged-in Users)
# =============================================================================

@login_required
def change_password_view(request):
    """
    Allow logged-in users to change their password.
    """
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        
        # Verify old password
        if not request.user.check_password(old_password):
            messages.error(request, 'Current password is incorrect.')
            return render(request, 'auth/change_password.html', {
                'page_title': 'Change Password'
            })
        
        # Validate new passwords
        if not password1 or not password2:
            messages.error(request, 'Please enter both new password fields.')
            return render(request, 'auth/change_password.html', {
                'page_title': 'Change Password'
            })
        
        if password1 != password2:
            messages.error(request, 'New passwords do not match.')
            return render(request, 'auth/change_password.html', {
                'page_title': 'Change Password'
            })
        
        if len(password1) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
            return render(request, 'auth/change_password.html', {
                'page_title': 'Change Password'
            })
        
        if old_password == password1:
            messages.error(request, 'New password must be different from current password.')
            return render(request, 'auth/change_password.html', {
                'page_title': 'Change Password'
            })
        
        # Set the new password
        request.user.set_password(password1)
        request.user.save()
        
        # Log the password change
        AuditLog.objects.create(
            user=request.user,
            action_type='UPDATE',
            model_name='User',
            object_id=request.user.id,
            object_repr=f'Password changed for {request.user.username}',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:300]
        )
        
        # Update session to prevent logout
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, request.user)
        
        messages.success(request, 'Your password has been changed successfully!')
        return redirect('dashboard')
    
    context = {
        'page_title': 'Change Password - MediSphere Hospital'
    }
    return render(request, 'auth/change_password.html', context)


# =============================================================================
# PASSWORD RESET VIA OTP (Alternative Method)
# =============================================================================

def forgot_password_otp_view(request):
    """
    Alternative password reset using OTP sent to phone.
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Step 1: Send OTP
        if action == 'send_otp':
            phone_number = request.POST.get('phone_number', '').strip()
            
            if not phone_number:
                messages.error(request, 'Please enter your phone number.')
                return render(request, 'auth/forgot_password_otp.html')
            
            # Find user by phone number
            try:
                user = User.objects.get(phone_number=phone_number, is_active=True)
            except User.DoesNotExist:
                messages.error(request, 'No account found with this phone number.')
                return render(request, 'auth/forgot_password_otp.html')
            
            # Generate OTP
            otp = ''.join(random.choices(string.digits, k=6))
            
            # Store OTP in session (in production, use Redis or database)
            request.session['reset_otp'] = otp
            request.session['reset_user_id'] = user.id
            request.session['otp_timestamp'] = timezone.now().isoformat()
            
            # Send OTP via SMS
            send_otp_sms(user, otp)
            
            messages.success(request, f'OTP has been sent to {phone_number}')
            
            context = {
                'show_otp_form': True,
                'phone_number': phone_number,
                'page_title': 'Verify OTP'
            }
            return render(request, 'auth/forgot_password_otp.html', context)
        
        # Step 2: Verify OTP
        elif action == 'verify_otp':
            entered_otp = request.POST.get('otp', '').strip()
            stored_otp = request.session.get('reset_otp')
            user_id = request.session.get('reset_user_id')
            
            if not stored_otp or not user_id:
                messages.error(request, 'Session expired. Please try again.')
                return redirect('forgot_password_otp')
            
            if entered_otp != stored_otp:
                messages.error(request, 'Invalid OTP. Please try again.')
                context = {
                    'show_otp_form': True,
                    'phone_number': request.POST.get('phone_number'),
                    'page_title': 'Verify OTP'
                }
                return render(request, 'auth/forgot_password_otp.html', context)
            
            # OTP verified, show password reset form
            context = {
                'show_password_form': True,
                'user_id': user_id,
                'page_title': 'Reset Password'
            }
            return render(request, 'auth/forgot_password_otp.html', context)
        
        # Step 3: Reset Password
        elif action == 'reset_password':
            user_id = request.session.get('reset_user_id')
            password1 = request.POST.get('password1')
            password2 = request.POST.get('password2')
            
            if not user_id:
                messages.error(request, 'Session expired. Please try again.')
                return redirect('forgot_password_otp')
            
            if password1 != password2:
                messages.error(request, 'Passwords do not match.')
                context = {
                    'show_password_form': True,
                    'user_id': user_id,
                    'page_title': 'Reset Password'
                }
                return render(request, 'auth/forgot_password_otp.html', context)
            
            if len(password1) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
                context = {
                    'show_password_form': True,
                    'user_id': user_id,
                    'page_title': 'Reset Password'
                }
                return render(request, 'auth/forgot_password_otp.html', context)
            
            # Reset password
            user = User.objects.get(id=user_id)
            user.set_password(password1)
            user.save()
            
            # Clear session
            request.session.pop('reset_otp', None)
            request.session.pop('reset_user_id', None)
            request.session.pop('otp_timestamp', None)
            
            # Log the action
            AuditLog.objects.create(
                user=user,
                action_type='UPDATE',
                model_name='User',
                object_id=user.id,
                object_repr=f'Password reset via OTP for {user.username}',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:300]
            )
            
            messages.success(request, 'Password reset successful! You can now login.')
            return redirect('login')
    
    context = {
        'page_title': 'Forgot Password - MediSphere Hospital'
    }
    return render(request, 'auth/forgot_password_otp.html', context)


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def send_password_reset_sms(user, reset_url):
    """Send password reset link via SMS"""
    try:
        from main_application.models import HospitalSettings
        settings = HospitalSettings.load()
        
        if settings.sms_enabled:
            message = f"MediSphere Hospital: Password reset link: {reset_url}"
            
            # Log SMS (actual sending would use SMS gateway API)
            SMSLog.objects.create(
                patient=None,  # This is for staff, not patients
                phone_number=user.phone_number,
                sms_type='GENERAL',
                message=message,
                status='SENT'
            )
            
            # TODO: Integrate with actual SMS gateway (Africa's Talking, Twilio, etc.)
            # Example:
            # send_sms_via_gateway(user.phone_number, message)
            
    except Exception as e:
        pass  # Fail silently for SMS


def send_otp_sms(user, otp):
    """Send OTP via SMS"""
    try:
        from main_application.models import HospitalSettings
        settings = HospitalSettings.load()
        
        if settings.sms_enabled:
            message = f"MediSphere Hospital: Your password reset OTP is: {otp}. Valid for 10 minutes."
            
            # Log SMS
            SMSLog.objects.create(
                patient=None,
                phone_number=user.phone_number,
                sms_type='GENERAL',
                message=message,
                status='SENT'
            )
            
            # TODO: Integrate with actual SMS gateway
            
    except Exception as e:
        pass  # Fail silently

@login_required
def dashboard_view(request):
    """
    Main dashboard that redirects users to their role-specific dashboard.
    """
    user = request.user
    
    # Superuser/Admin dashboard
    if user.is_superuser:
        return redirect('admin_dashboard')
    
    # Role-based dashboard routing
    if user.role:
        role_name = user.role.name
        
        # Medical Staff
        if role_name == 'MEDICAL_SUPERINTENDENT':
            return redirect('superintendent_dashboard')
        elif role_name in ['DOCTOR', 'CLINICAL_OFFICER']:
            return redirect('doctor_dashboard')
        elif role_name == 'NURSE':
            return redirect('nurse_dashboard')
        
        # Diagnostic Services
        elif role_name == 'LAB_TECHNICIAN':
            return redirect('lab_dashboard')
        elif role_name == 'RADIOLOGIST':
            return redirect('radiology_dashboard')
        
        # Pharmacy
        elif role_name == 'PHARMACIST':
            return redirect('pharmacy_dashboard')
        
        # Administrative
        elif role_name == 'RECEPTIONIST':
            return redirect('reception_dashboard')
        elif role_name == 'CASHIER':
            return redirect('billing_dashboard')
        elif role_name == 'NHIF_OFFICER':
            return redirect('nhif_dashboard')
        elif role_name == 'IT_ADMIN':
            return redirect('admin_dashboard')
    
    # Default fallback dashboard
    return redirect('default_dashboard')


# =============================================================================
# ROLE-SPECIFIC DASHBOARD VIEWS (Placeholders)
# =============================================================================

@login_required
@role_required('MEDICAL_SUPERINTENDENT')
def superintendent_dashboard(request):
    """Medical Superintendent Dashboard"""
    context = {
        'page_title': 'Medical Superintendent Dashboard',
        'user': request.user
    }
    return render(request, 'dashboards/superintendent.html', context)


@login_required
@role_required('DOCTOR', 'CLINICAL_OFFICER')
def doctor_dashboard(request):
    """Doctor/Clinical Officer Dashboard"""
    context = {
        'page_title': 'Doctor Dashboard',
        'user': request.user
    }
    return render(request, 'dashboards/doctor.html', context)


@login_required
@role_required('NURSE')
def nurse_dashboard(request):
    """Nurse Dashboard"""
    context = {
        'page_title': 'Nurse Dashboard',
        'user': request.user
    }
    return render(request, 'dashboards/nurse.html', context)


@login_required
@role_required('LAB_TECHNICIAN')
def lab_dashboard(request):
    """Laboratory Dashboard"""
    context = {
        'page_title': 'Laboratory Dashboard',
        'user': request.user
    }
    return render(request, 'dashboards/lab.html', context)


@login_required
@role_required('RADIOLOGIST')
def radiology_dashboard(request):
    """Radiology Dashboard"""
    context = {
        'page_title': 'Radiology Dashboard',
        'user': request.user
    }
    return render(request, 'dashboards/radiology.html', context)


@login_required
@role_required('PHARMACIST')
def pharmacy_dashboard(request):
    """Pharmacy Dashboard"""
    context = {
        'page_title': 'Pharmacy Dashboard',
        'user': request.user
    }
    return render(request, 'dashboards/pharmacy.html', context)


# @login_required
# @role_required('RECEPTIONIST')
# def reception_dashboard(request):
#     """Reception Dashboard"""
#     context = {
#         'page_title': 'Reception Dashboard',
#         'user': request.user
#     }
#     return render(request, 'dashboards/reception.html', context)


@login_required
@role_required('CASHIER')
def billing_dashboard(request):
    """Billing/Cashier Dashboard"""
    context = {
        'page_title': 'Billing Dashboard',
        'user': request.user
    }
    return render(request, 'dashboards/billing.html', context)


@login_required
@role_required('NHIF_OFFICER')
def nhif_dashboard(request):
    """NHIF Officer Dashboard"""
    context = {
        'page_title': 'NHIF Dashboard',
        'user': request.user
    }
    return render(request, 'dashboards/nhif.html', context)


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta, datetime
import json
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    Patient, PatientVisit, Consultation, Invoice, Payment,
    Drug, DrugStock, LabOrder, RadiologyOrder, Admission,
    User, Department, NHIFClaim, Prescription, Surgery
)


@login_required
def admin_dashboard(request):
    """Enhanced Admin/IT Dashboard with Analytics"""
    if not (request.user.is_superuser or (request.user.role and request.user.role.name == 'IT_ADMIN')):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # Get filter parameters
    period = request.GET.get('period', '30')  # Default 30 days
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Calculate date range
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date = timezone.now().date() - timedelta(days=int(period))
            end_date = timezone.now().date()
    else:
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=int(period))
    
    # ========== KEY STATISTICS ==========
    total_patients = Patient.objects.filter(is_active=True).count()
    total_staff = User.objects.filter(is_active_staff=True).count()
    
    # Period-specific stats
    visits_in_period = PatientVisit.objects.filter(
        visit_date__gte=start_date,
        visit_date__lte=end_date
    )
    total_visits = visits_in_period.count()
    
    # Revenue statistics
    revenue_data = Invoice.objects.filter(
        invoice_date__date__gte=start_date,
        invoice_date__date__lte=end_date
    ).aggregate(
        total_revenue=Sum('total_amount'),
        total_paid=Sum('amount_paid'),
        total_pending=Sum('balance')
    )
    
    total_revenue = revenue_data['total_revenue'] or Decimal('0.00')
    total_paid = revenue_data['total_paid'] or Decimal('0.00')
    total_pending = revenue_data['total_pending'] or Decimal('0.00')
    
    # Payment methods breakdown
    payment_methods = Payment.objects.filter(
        payment_date__date__gte=start_date,
        payment_date__date__lte=end_date,
        status='COMPLETED'
    ).values('payment_method').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    payment_labels = [pm['payment_method'] for pm in payment_methods]
    payment_totals = [float(pm['total']) for pm in payment_methods]
    
    # ========== PATIENT ANALYTICS ==========
    
    # Patient registrations trend (last 30 days)
    registration_trend = Patient.objects.filter(
        registration_date__date__gte=start_date,
        registration_date__date__lte=end_date
    ).annotate(
        date=TruncDate('registration_date')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')
    
    reg_dates = [rt['date'].strftime('%Y-%m-%d') for rt in registration_trend]
    reg_counts = [rt['count'] for rt in registration_trend]
    
    # Patient demographics - Age groups (Database-agnostic approach)
    from collections import Counter
    
    active_patients = Patient.objects.filter(is_active=True)
    age_group_counter = Counter()
    
    for patient in active_patients:
        age = patient.age
        if age < 1:
            age_group_counter['Infant (0-1)'] += 1
        elif age < 13:
            age_group_counter['Child (1-12)'] += 1
        elif age < 18:
            age_group_counter['Teen (13-17)'] += 1
        elif age < 65:
            age_group_counter['Adult (18-64)'] += 1
        else:
            age_group_counter['Elderly (65+)'] += 1
    
    age_group_labels = list(age_group_counter.keys())
    age_group_counts = list(age_group_counter.values())
    
    # Gender distribution
    gender_dist = Patient.objects.filter(is_active=True).values('gender').annotate(
        count=Count('id')
    )
    gender_labels = [g['gender'] for g in gender_dist]
    gender_counts = [g['count'] for g in gender_dist]
    
    # ========== VISIT ANALYTICS ==========
    
    # Daily visits trend
    daily_visits = visits_in_period.annotate(
        date=TruncDate('visit_date')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')
    
    visit_dates = [dv['date'].strftime('%Y-%m-%d') for dv in daily_visits]
    visit_counts = [dv['count'] for dv in daily_visits]
    
    # Visit types distribution
    visit_types = visits_in_period.values('visit_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    visit_type_labels = [vt['visit_type'] for vt in visit_types]
    visit_type_counts = [vt['count'] for vt in visit_types]
    
    # Visit status distribution
    visit_status = visits_in_period.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    visit_status_labels = [vs['status'] for vs in visit_status]
    visit_status_counts = [vs['count'] for vs in visit_status]
    
    # Average wait time
    completed_visits = visits_in_period.filter(exit_time__isnull=False)
    if completed_visits.exists():
        avg_wait_time = sum([v.wait_time_minutes for v in completed_visits]) / completed_visits.count()
    else:
        avg_wait_time = 0
    
    # ========== REVENUE ANALYTICS ==========
    
    # Daily revenue trend
    daily_revenue = Invoice.objects.filter(
        invoice_date__date__gte=start_date,
        invoice_date__date__lte=end_date
    ).annotate(
        date=TruncDate('invoice_date')
    ).values('date').annotate(
        revenue=Sum('total_amount'),
        paid=Sum('amount_paid')
    ).order_by('date')
    
    revenue_dates = [dr['date'].strftime('%Y-%m-%d') for dr in daily_revenue]
    revenue_amounts = [float(dr['revenue'] or 0) for dr in daily_revenue]
    revenue_paid = [float(dr['paid'] or 0) for dr in daily_revenue]
    
    # Revenue by service type
    revenue_by_service = Invoice.objects.filter(
        invoice_date__date__gte=start_date,
        invoice_date__date__lte=end_date
    ).values('items__item_type').annotate(
        total=Sum(F('items__quantity') * F('items__unit_price'))
    ).order_by('-total')
    
    service_labels = [rs['items__item_type'] or 'OTHER' for rs in revenue_by_service]
    service_amounts = [float(rs['total'] or 0) for rs in revenue_by_service]
    
    # Top revenue generating departments
    top_departments = Consultation.objects.filter(
        consultation_start__date__gte=start_date,
        consultation_start__date__lte=end_date
    ).values('doctor__department__name').annotate(
        revenue=Sum('consultation_fee'),
        count=Count('id')
    ).order_by('-revenue')[:5]
    
    dept_labels = [td['doctor__department__name'] or 'N/A' for td in top_departments]
    dept_revenue = [float(td['revenue'] or 0) for td in top_departments]
    
    # ========== PHARMACY ANALYTICS ==========
    
    # Low stock medications
    low_stock = Drug.objects.filter(is_active=True).annotate(
        current_stock=Sum('stock_records__quantity')
    ).filter(
        current_stock__lte=F('reorder_level')
    ).order_by('current_stock')[:10]
    
    # Top selling medications
    top_medicines = Prescription.objects.filter(
        prescribed_at__date__gte=start_date,
        prescribed_at__date__lte=end_date
    ).values('items__drug__name').annotate(
        quantity=Sum('items__dispensed_quantity'),
        revenue=Sum(F('items__dispensed_quantity') * F('items__drug__unit_price'))
    ).order_by('-quantity')[:10]
    
    medicine_labels = [tm['items__drug__name'] for tm in top_medicines]
    medicine_quantities = [tm['quantity'] or 0 for tm in top_medicines]
    medicine_revenue = [float(tm['revenue'] or 0) for tm in top_medicines]
    
    # Expiring stock (next 90 days)
    expiring_soon = DrugStock.objects.filter(
        expiry_date__lte=timezone.now().date() + timedelta(days=90),
        expiry_date__gt=timezone.now().date(),
        quantity__gt=0
    ).select_related('drug').order_by('expiry_date')[:10]
    
    # ========== LAB & RADIOLOGY ANALYTICS ==========
    
    # Lab tests ordered
    lab_tests = LabOrder.objects.filter(
        ordered_at__date__gte=start_date,
        ordered_at__date__lte=end_date
    ).values('test__category').annotate(
        count=Count('id')
    ).order_by('-count')
    
    lab_labels = [lt['test__category'] for lt in lab_tests]
    lab_counts = [lt['count'] for lt in lab_tests]
    
    # Lab turnaround time
    completed_labs = LabOrder.objects.filter(
        ordered_at__date__gte=start_date,
        ordered_at__date__lte=end_date,
        status='COMPLETED',
        result__isnull=False
    ).select_related('result')
    
    if completed_labs.exists():
        total_tat = sum([
            (lab.result.result_date - lab.ordered_at).total_seconds() / 3600 
            for lab in completed_labs
        ])
        avg_lab_tat = total_tat / completed_labs.count()
    else:
        avg_lab_tat = 0
    
    # Radiology tests
    radiology_tests = RadiologyOrder.objects.filter(
        ordered_at__date__gte=start_date,
        ordered_at__date__lte=end_date
    ).values('test__modality').annotate(
        count=Count('id')
    ).order_by('-count')
    
    rad_labels = [rt['test__modality'] for rt in radiology_tests]
    rad_counts = [rt['count'] for rt in radiology_tests]
    
    # ========== INPATIENT ANALYTICS ==========
    
    # Active admissions
    active_admissions = Admission.objects.filter(status='ACTIVE').count()
    
    # Bed occupancy by ward
    from .models import Ward
    ward_occupancy = Ward.objects.filter(is_active=True).annotate(
        occupied=Count('beds', filter=Q(beds__is_occupied=True)),
        available=Count('beds', filter=Q(beds__is_occupied=False, beds__is_available=True))
    )
    
    ward_labels = [w.name for w in ward_occupancy]
    ward_occupied = [w.occupied for w in ward_occupancy]
    ward_available = [w.available for w in ward_occupancy]
    
    # Average length of stay
    discharged_admissions = Admission.objects.filter(
        admission_datetime__date__gte=start_date,
        discharge_datetime__date__lte=end_date,
        status='DISCHARGED'
    )
    
    if discharged_admissions.exists():
        avg_los = sum([adm.length_of_stay for adm in discharged_admissions]) / discharged_admissions.count()
    else:
        avg_los = 0
    
    # ========== NHIF ANALYTICS ==========
    
    # NHIF claims status
    nhif_claims = NHIFClaim.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).values('status').annotate(
        count=Count('id'),
        amount=Sum('claimed_amount')
    )
    
    nhif_status_labels = [nc['status'] for nc in nhif_claims]
    nhif_status_counts = [nc['count'] for nc in nhif_claims]
    nhif_status_amounts = [float(nc['amount'] or 0) for nc in nhif_claims]
    
    # ========== STAFF PERFORMANCE ==========
    
    # Top performing doctors by consultations
    top_doctors = Consultation.objects.filter(
        consultation_start__date__gte=start_date,
        consultation_start__date__lte=end_date
    ).values('doctor__first_name', 'doctor__last_name').annotate(
        count=Count('id'),
        revenue=Sum('consultation_fee')
    ).order_by('-count')[:10]
    
    doctor_labels = [f"Dr. {td['doctor__first_name']} {td['doctor__last_name']}" for td in top_doctors]
    doctor_counts = [td['count'] for td in top_doctors]
    doctor_revenue = [float(td['revenue'] or 0) for td in top_doctors]
    
    # Department performance
    dept_performance = Department.objects.filter(is_active=True).annotate(
        staff_count=Count('staff', filter=Q(staff__is_active_staff=True)),
        visits=Count('staff__consultations', filter=Q(
            staff__consultations__consultation_start__date__gte=start_date,
            staff__consultations__consultation_start__date__lte=end_date
        ))
    ).order_by('-visits')
    
    # ========== SURGICAL ANALYTICS ==========
    
    surgeries_performed = Surgery.objects.filter(
        start_time__date__gte=start_date,
        start_time__date__lte=end_date,
        status='COMPLETED'
    ).count()
    
    surgery_types = Surgery.objects.filter(
        start_time__date__gte=start_date,
        start_time__date__lte=end_date
    ).values('surgery_type').annotate(
        count=Count('id')
    )
    
    surgery_type_labels = [st['surgery_type'] for st in surgery_types]
    surgery_type_counts = [st['count'] for st in surgery_types]
    
    # ========== HANDLE EXCEL EXPORT ==========
    if request.GET.get('export') == 'excel':
        return export_dashboard_to_excel(
            start_date, end_date, total_patients, total_staff, total_visits,
            total_revenue, total_paid, total_pending, active_admissions,
            surgeries_performed, low_stock, expiring_soon, visit_types,
            payment_methods, top_medicines, top_doctors, ward_occupancy
        )
    
    context = {
        'page_title': 'Admin Dashboard',
        'user': request.user,
        
        # Filter parameters
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        
        # Key Statistics
        'total_patients': total_patients,
        'total_staff': total_staff,
        'total_visits': total_visits,
        'total_revenue': total_revenue,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'active_admissions': active_admissions,
        'surgeries_performed': surgeries_performed,
        'avg_wait_time': round(avg_wait_time, 1),
        'avg_lab_tat': round(avg_lab_tat, 1),
        'avg_los': round(avg_los, 1),
        
        # Patient Analytics
        'reg_dates': json.dumps(reg_dates),
        'reg_counts': json.dumps(reg_counts),
        'age_group_labels': json.dumps(age_group_labels),
        'age_group_counts': json.dumps(age_group_counts),
        'gender_labels': json.dumps(gender_labels),
        'gender_counts': json.dumps(gender_counts),
        
        # Visit Analytics
        'visit_dates': json.dumps(visit_dates),
        'visit_counts': json.dumps(visit_counts),
        'visit_type_labels': json.dumps(visit_type_labels),
        'visit_type_counts': json.dumps(visit_type_counts),
        'visit_status_labels': json.dumps(visit_status_labels),
        'visit_status_counts': json.dumps(visit_status_counts),
        
        # Revenue Analytics
        'revenue_dates': json.dumps(revenue_dates),
        'revenue_amounts': json.dumps(revenue_amounts),
        'revenue_paid': json.dumps(revenue_paid),
        'payment_labels': json.dumps(payment_labels),
        'payment_totals': json.dumps(payment_totals),
        'service_labels': json.dumps(service_labels),
        'service_amounts': json.dumps(service_amounts),
        'dept_labels': json.dumps(dept_labels),
        'dept_revenue': json.dumps(dept_revenue),
        
        # Pharmacy Analytics
        'low_stock': low_stock,
        'expiring_soon': expiring_soon,
        'medicine_labels': json.dumps(medicine_labels),
        'medicine_quantities': json.dumps(medicine_quantities),
        'medicine_revenue': json.dumps(medicine_revenue),
        
        # Lab & Radiology
        'lab_labels': json.dumps(lab_labels),
        'lab_counts': json.dumps(lab_counts),
        'rad_labels': json.dumps(rad_labels),
        'rad_counts': json.dumps(rad_counts),
        
        # Inpatient Analytics
        'ward_labels': json.dumps(ward_labels),
        'ward_occupied': json.dumps(ward_occupied),
        'ward_available': json.dumps(ward_available),
        
        # NHIF Analytics
        'nhif_status_labels': json.dumps(nhif_status_labels),
        'nhif_status_counts': json.dumps(nhif_status_counts),
        'nhif_status_amounts': json.dumps(nhif_status_amounts),
        
        # Staff Performance
        'doctor_labels': json.dumps(doctor_labels),
        'doctor_counts': json.dumps(doctor_counts),
        'doctor_revenue': json.dumps(doctor_revenue),
        'dept_performance': dept_performance,
        
        # Surgery Analytics
        'surgery_type_labels': json.dumps(surgery_type_labels),
        'surgery_type_counts': json.dumps(surgery_type_counts),
    }
    
    return render(request, 'dashboards/admin_dashboard.html', context)


def export_dashboard_to_excel(start_date, end_date, total_patients, total_staff, 
                               total_visits, total_revenue, total_paid, total_pending,
                               active_admissions, surgeries_performed, low_stock, 
                               expiring_soon, visit_types, payment_methods, 
                               top_medicines, top_doctors, ward_occupancy):
    """Export dashboard data to Excel"""
    
    wb = openpyxl.Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    title_font = Font(bold=True, size=14, color='2980b9')
    
    # ========== SUMMARY SHEET ==========
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    
    # Title
    ws_summary['A1'] = 'MediSphere Hospital - Dashboard Report'
    ws_summary['A1'].font = title_font
    ws_summary['A2'] = f'Period: {start_date} to {end_date}'
    
    # Key Metrics
    ws_summary['A4'] = 'Key Performance Indicators'
    ws_summary['A4'].font = Font(bold=True, size=12)
    
    metrics = [
        ('Total Patients', total_patients),
        ('Total Staff', total_staff),
        ('Total Visits', total_visits),
        ('Total Revenue (KES)', float(total_revenue)),
        ('Amount Paid (KES)', float(total_paid)),
        ('Pending Amount (KES)', float(total_pending)),
        ('Active Admissions', active_admissions),
        ('Surgeries Performed', surgeries_performed),
    ]
    
    row = 5
    for metric, value in metrics:
        ws_summary[f'A{row}'] = metric
        ws_summary[f'B{row}'] = value
        row += 1
    
    # ========== VISIT TYPES SHEET ==========
    ws_visits = wb.create_sheet('Visit Types')
    ws_visits['A1'] = 'Visit Type'
    ws_visits['B1'] = 'Count'
    ws_visits['A1'].fill = header_fill
    ws_visits['B1'].fill = header_fill
    ws_visits['A1'].font = header_font
    ws_visits['B1'].font = header_font
    
    row = 2
    for vt in visit_types:
        ws_visits[f'A{row}'] = vt['visit_type']
        ws_visits[f'B{row}'] = vt['count']
        row += 1
    
    # ========== PAYMENT METHODS SHEET ==========
    ws_payments = wb.create_sheet('Payment Methods')
    ws_payments['A1'] = 'Payment Method'
    ws_payments['B1'] = 'Total Amount (KES)'
    ws_payments['C1'] = 'Count'
    for cell in ['A1', 'B1', 'C1']:
        ws_payments[cell].fill = header_fill
        ws_payments[cell].font = header_font
    
    row = 2
    for pm in payment_methods:
        ws_payments[f'A{row}'] = pm['payment_method']
        ws_payments[f'B{row}'] = float(pm['total'])
        ws_payments[f'C{row}'] = pm['count']
        row += 1
    
    # ========== TOP MEDICINES SHEET ==========
    ws_meds = wb.create_sheet('Top Medicines')
    ws_meds['A1'] = 'Medicine'
    ws_meds['B1'] = 'Quantity Dispensed'
    ws_meds['C1'] = 'Revenue (KES)'
    for cell in ['A1', 'B1', 'C1']:
        ws_meds[cell].fill = header_fill
        ws_meds[cell].font = header_font
    
    row = 2
    for med in top_medicines:
        ws_meds[f'A{row}'] = med['items__drug__name']
        ws_meds[f'B{row}'] = med['quantity'] or 0
        ws_meds[f'C{row}'] = float(med['revenue'] or 0)
        row += 1
    
    # ========== LOW STOCK SHEET ==========
    ws_stock = wb.create_sheet('Low Stock')
    ws_stock['A1'] = 'Medicine'
    ws_stock['B1'] = 'Current Stock'
    ws_stock['C1'] = 'Reorder Level'
    for cell in ['A1', 'B1', 'C1']:
        ws_stock[cell].fill = header_fill
        ws_stock[cell].font = header_font
    
    row = 2
    for drug in low_stock:
        ws_stock[f'A{row}'] = drug.name
        ws_stock[f'B{row}'] = drug.current_stock
        ws_stock[f'C{row}'] = drug.reorder_level
        row += 1
    
    # ========== EXPIRING STOCK SHEET ==========
    ws_expiry = wb.create_sheet('Expiring Stock')
    ws_expiry['A1'] = 'Medicine'
    ws_expiry['B1'] = 'Batch Number'
    ws_expiry['C1'] = 'Quantity'
    ws_expiry['D1'] = 'Expiry Date'
    ws_expiry['E1'] = 'Days to Expiry'
    for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
        ws_expiry[cell].fill = header_fill
        ws_expiry[cell].font = header_font
    
    row = 2
    for stock in expiring_soon:
        ws_expiry[f'A{row}'] = stock.drug.name
        ws_expiry[f'B{row}'] = stock.batch_number
        ws_expiry[f'C{row}'] = stock.quantity
        ws_expiry[f'D{row}'] = stock.expiry_date.strftime('%Y-%m-%d')
        ws_expiry[f'E{row}'] = stock.days_to_expiry
        row += 1
    
    # ========== TOP DOCTORS SHEET ==========
    ws_doctors = wb.create_sheet('Top Doctors')
    ws_doctors['A1'] = 'Doctor'
    ws_doctors['B1'] = 'Consultations'
    ws_doctors['C1'] = 'Revenue (KES)'
    for cell in ['A1', 'B1', 'C1']:
        ws_doctors[cell].fill = header_fill
        ws_doctors[cell].font = header_font
    
    row = 2
    for doc in top_doctors:
        ws_doctors[f'A{row}'] = f"Dr. {doc['doctor__first_name']} {doc['doctor__last_name']}"
        ws_doctors[f'B{row}'] = doc['count']
        ws_doctors[f'C{row}'] = float(doc['revenue'] or 0)
        row += 1
    
    # ========== WARD OCCUPANCY SHEET ==========
    ws_wards = wb.create_sheet('Ward Occupancy')
    ws_wards['A1'] = 'Ward'
    ws_wards['B1'] = 'Occupied Beds'
    ws_wards['C1'] = 'Available Beds'
    ws_wards['D1'] = 'Total Beds'
    ws_wards['E1'] = 'Occupancy Rate (%)'
    for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
        ws_wards[cell].fill = header_fill
        ws_wards[cell].font = header_font
    
    row = 2
    for ward in ward_occupancy:
        ws_wards[f'A{row}'] = ward.name
        ws_wards[f'B{row}'] = ward.occupied
        ws_wards[f'C{row}'] = ward.available
        ws_wards[f'D{row}'] = ward.total_beds
        ws_wards[f'E{row}'] = ward.occupancy_rate
        row += 1
    
    # Auto-adjust column widths for all sheets
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=dashboard_report_{start_date}_{end_date}.xlsx'
    
    wb.save(response)
    return response


@login_required
def default_dashboard(request):
    """Default dashboard for users without specific roles"""
    context = {
        'page_title': 'Dashboard',
        'user': request.user
    }
    return render(request, 'dashboards/default.html', context)


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def error_404(request, exception):
    return render(request, 'errors/404.html', status=404)

def error_500(request):
    return render(request, 'errors/500.html', status=500)

def error_403(request, exception):
    return render(request, 'errors/403.html', status=403)

def error_400(request, exception):
    return render(request, 'errors/400.html', status=400)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import csv
from datetime import datetime

from .models import Patient, User
from .forms import PatientRegistrationForm  # You'll need to create this


# =============================================================================
# PATIENT LIST VIEW (with Search & Export)
# =============================================================================

@login_required
def patients_list(request):
    """
    Display all patients with search functionality and pagination
    """
    # Get search query
    search_query = request.GET.get('search', '').strip()
    
    # Base queryset
    patients = Patient.objects.filter(is_active=True).select_related('registered_by')
    
    # Apply search filters
    if search_query:
        patients = patients.filter(
            Q(patient_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(middle_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id_number__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(nhif_number__icontains=search_query)
        )
    
    # Order by most recent
    patients = patients.order_by('-registration_date')
    
    # Handle export requests
    export_format = request.GET.get('export')
    if export_format == 'excel':
        return export_patients_excel(patients, search_query)
    elif export_format == 'csv':
        return export_patients_csv(patients, search_query)
    
    # Pagination
    paginator = Paginator(patients, 25)  # 25 patients per page
    page_number = request.GET.get('page', 1)
    patients_page = paginator.get_page(page_number)
    
    context = {
        'patients': patients_page,
        'search_query': search_query,
        'total_patients': patients.count(),
        'page_title': 'Patient Records',
    }
    
    return render(request, 'patients/patients_list.html', context)


# =============================================================================
# PATIENT REGISTRATION
# =============================================================================

@login_required
def patients_create(request):
    """
    Register a new patient
    """
    if request.method == 'POST':
        form = PatientRegistrationForm(request.POST)
        
        if form.is_valid():
            patient = form.save(commit=False)
            patient.registered_by = request.user
            patient.save()
            
            messages.success(
                request, 
                f'Patient {patient.full_name} successfully registered with number {patient.patient_number}!'
            )
            return redirect('patients-detail', patient_number=patient.patient_number)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PatientRegistrationForm()
    
    context = {
        'form': form,
        'page_title': 'New Patient Registration',
    }
    
    return render(request, 'patients/patients_create.html', context)


# =============================================================================
# PATIENT DETAIL VIEW
# =============================================================================

@login_required
def patients_detail(request, patient_number):
    """
    View detailed patient information using patient number
    """
    patient = get_object_or_404(
        Patient.objects.select_related('registered_by'),
        patient_number=patient_number,
        is_active=True
    )
    
    # Get patient's recent visits
    recent_visits = patient.visits.all().order_by('-visit_date')[:10]
    
    # Get patient's prescriptions
    recent_prescriptions = patient.visits.filter(
        prescriptions__isnull=False
    ).order_by('-visit_date')[:5]
    
    # Get patient's invoices
    recent_invoices = patient.invoices.all().order_by('-invoice_date')[:10]
    
    # Get patient's admissions
    admissions = patient.admissions.all().order_by('-admission_datetime')[:5]
    
    # Calculate statistics
    total_visits = patient.visits.count()
    total_amount_billed = sum(invoice.total_amount for invoice in patient.invoices.all())
    total_amount_paid = sum(invoice.amount_paid for invoice in patient.invoices.all())
    outstanding_balance = sum(invoice.balance for invoice in patient.invoices.all())
    
    context = {
        'patient': patient,
        'recent_visits': recent_visits,
        'recent_prescriptions': recent_prescriptions,
        'recent_invoices': recent_invoices,
        'admissions': admissions,
        'total_visits': total_visits,
        'total_amount_billed': total_amount_billed,
        'total_amount_paid': total_amount_paid,
        'outstanding_balance': outstanding_balance,
        'page_title': f'Patient: {patient.full_name}',
    }
    
    return render(request, 'patients/patients_detail.html', context)


# =============================================================================
# PATIENT UPDATE
# =============================================================================

@login_required
def patients_update(request, patient_number):
    """
    Update patient information
    """
    patient = get_object_or_404(Patient, patient_number=patient_number, is_active=True)
    
    if request.method == 'POST':
        form = PatientRegistrationForm(request.POST, instance=patient)
        
        if form.is_valid():
            patient = form.save()
            messages.success(request, f'Patient {patient.full_name} updated successfully!')
            return redirect('patients-detail', patient_number=patient.patient_number)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PatientRegistrationForm(instance=patient)
    
    context = {
        'form': form,
        'patient': patient,
        'page_title': f'Edit Patient: {patient.full_name}',
    }
    
    return render(request, 'patients/patients_update.html', context)


# =============================================================================
# PATIENT DELETE (Soft Delete)
# =============================================================================

@login_required
@require_http_methods(["POST"])
def patients_delete(request, patient_number):
    """
    Soft delete a patient (mark as inactive)
    Note: This doesn't actually delete from database to maintain data integrity
    """
    patient = get_object_or_404(Patient, patient_number=patient_number)
    
    # Check if patient has any visits
    if patient.visits.exists():
        messages.warning(
            request, 
            f'Cannot delete patient {patient.full_name} as they have visit records. Patient marked as inactive instead.'
        )
        patient.is_active = False
        patient.save()
    else:
        # If no visits, can safely soft delete
        patient.is_active = False
        patient.notes += f"\n\nDeactivated by {request.user.get_full_name()} on {timezone.now()}"
        patient.save()
        messages.success(request, f'Patient {patient.full_name} has been deactivated.')
    
    return redirect('patients-list')





# =============================================================================
# PATIENT SEARCH API (Autocomplete)
# =============================================================================

@login_required
def patient_search_api(request):
    """
    API endpoint for patient autocomplete search
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    patients = Patient.objects.filter(
        Q(patient_number__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(id_number__icontains=query) |
        Q(phone_number__icontains=query),
        is_active=True
    )[:10]
    
    results = [{
        'id': p.id,
        'patient_number': p.patient_number,
        'name': p.full_name,
        'id_number': p.id_number,
        'phone': p.phone_number,
        'age': p.age,
        'gender': p.get_gender_display(),
    } for p in patients]
    
    return JsonResponse({'results': results})


# =============================================================================
# EXPORT FUNCTIONS
# =============================================================================

def export_patients_excel(queryset, search_query=''):
    """
    Export patients to Excel file
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Patients'
    
    # Define styles
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    title_font = Font(bold=True, size=14, color='2980b9')
    
    # Title
    ws['A1'] = 'Patient Records Export'
    ws['A1'].font = title_font
    ws['A2'] = f'Generated: {timezone.now().strftime("%B %d, %Y %H:%M")}'
    if search_query:
        ws['A3'] = f'Search Query: {search_query}'
    ws['A4'] = f'Total Records: {queryset.count()}'
    
    # Headers
    headers = [
        'Patient Number', 'Full Name', 'ID Number', 'Gender', 'Date of Birth', 
        'Age', 'Phone Number', 'Email', 'County', 'Blood Group', 
        'NHIF Status', 'NHIF Number', 'Registration Date', 'Registered By'
    ]
    
    start_row = 6
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, patient in enumerate(queryset, start_row + 1):
        ws.cell(row=row_num, column=1, value=patient.patient_number)
        ws.cell(row=row_num, column=2, value=patient.full_name)
        ws.cell(row=row_num, column=3, value=patient.id_number or 'N/A')
        ws.cell(row=row_num, column=4, value=patient.get_gender_display())
        ws.cell(row=row_num, column=5, value=patient.date_of_birth.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=6, value=patient.age)
        ws.cell(row=row_num, column=7, value=patient.phone_number)
        ws.cell(row=row_num, column=8, value=patient.email or 'N/A')
        ws.cell(row=row_num, column=9, value=patient.county)
        ws.cell(row=row_num, column=10, value=patient.blood_group or 'N/A')
        ws.cell(row=row_num, column=11, value=patient.get_nhif_status_display())
        ws.cell(row=row_num, column=12, value=patient.nhif_number or 'N/A')
        ws.cell(row=row_num, column=13, value=patient.registration_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=14, value=patient.registered_by.get_full_name())
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def export_patients_csv(queryset, search_query=''):
    """
    Export patients to CSV file
    """
    response = HttpResponse(content_type='text/csv')
    filename = f'patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([
        'Patient Number', 'Full Name', 'ID Number', 'Gender', 'Date of Birth',
        'Age', 'Phone Number', 'Email', 'County', 'Blood Group',
        'NHIF Status', 'NHIF Number', 'Registration Date', 'Registered By'
    ])
    
    # Write data rows
    for patient in queryset:
        writer.writerow([
            patient.patient_number,
            patient.full_name,
            patient.id_number or 'N/A',
            patient.get_gender_display(),
            patient.date_of_birth.strftime('%Y-%m-%d'),
            patient.age,
            patient.phone_number,
            patient.email or 'N/A',
            patient.county,
            patient.blood_group or 'N/A',
            patient.get_nhif_status_display(),
            patient.nhif_number or 'N/A',
            patient.registration_date.strftime('%Y-%m-%d'),
            patient.registered_by.get_full_name()
        ])
    
    return response

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg
from django.http import JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta

from .models import (
    PatientVisit, Patient, TriageAssessment, 
    Consultation, User, Invoice
)
from .forms import PatientVisitForm, TriageAssessmentForm


# =============================================================================
# REGISTER VISIT
# =============================================================================

@login_required
def visits_register(request):
    """
    Register a new patient visit
    """
    if request.method == 'POST':
        form = PatientVisitForm(request.POST)
        
        if form.is_valid():
            visit = form.save()
            
            messages.success(
                request,
                f'Visit {visit.visit_number} registered successfully for {visit.patient.full_name}!'
            )
            
            # Redirect based on visit type
            if visit.visit_type in ['EMERGENCY', 'AMBULANCE']:
                return redirect('triage-assessment', visit_number=visit.visit_number)
            else:
                return redirect('visits-detail', visit_number=visit.visit_number)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PatientVisitForm()
    
    # Get recent patients for quick selection
    recent_patients = Patient.objects.filter(is_active=True).order_by('-registration_date')[:10]
    
    context = {
        'form': form,
        'recent_patients': recent_patients,
        'page_title': 'Register Patient Visit',
    }
    
    return render(request, 'visits/visits_register.html', context)


# =============================================================================
# TRIAGE QUEUE
# =============================================================================

@login_required
def triage_queue(request):
    """
    Display triage queue - patients waiting for triage assessment
    """
    # Get visits awaiting triage
    pending_triage = PatientVisit.objects.filter(
        status__in=['WAITING', 'TRIAGE'],
        visit_date=timezone.now().date()
    ).select_related('patient').order_by('priority_level', 'arrival_time')
    
    # Get completed triages today
    completed_today = TriageAssessment.objects.filter(
        assessment_time__date=timezone.now().date()
    ).select_related('visit__patient', 'nurse').order_by('-assessment_time')[:10]
    
    # Statistics
    total_waiting = pending_triage.filter(status='WAITING').count()
    in_triage = pending_triage.filter(status='TRIAGE').count()
    
    # Critical cases
    critical_cases = pending_triage.filter(priority_level__lte=2)
    
    context = {
        'pending_triage': pending_triage,
        'completed_today': completed_today,
        'total_waiting': total_waiting,
        'in_triage': in_triage,
        'critical_cases': critical_cases,
        'page_title': 'Triage Queue',
    }
    
    return render(request, 'visits/triage_queue.html', context)


@login_required
def triage_assessment(request, visit_number):
    """
    Perform triage assessment
    """
    visit = get_object_or_404(
        PatientVisit.objects.select_related('patient'),
        visit_number=visit_number
    )
    
    # Check if triage already exists
    try:
        triage = visit.triage
        is_update = True
    except TriageAssessment.DoesNotExist:
        triage = None
        is_update = False
    
    if request.method == 'POST':
        if is_update:
            form = TriageAssessmentForm(request.POST, instance=triage)
        else:
            form = TriageAssessmentForm(request.POST)
        
        if form.is_valid():
            triage = form.save(commit=False)
            triage.visit = visit
            triage.nurse = request.user
            triage.save()
            
            # Update visit status
            visit.status = 'CONSULTATION'
            visit.priority_level = {
                'CRITICAL': 1,
                'EMERGENCY': 2,
                'URGENT': 3,
                'NORMAL': 4,
            }.get(triage.emergency_level, 4)
            visit.save()
            
            messages.success(
                request,
                f'Triage assessment completed for {visit.patient.full_name}. Patient moved to consultation queue.'
            )
            return redirect('triage-queue')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        if is_update:
            form = TriageAssessmentForm(instance=triage)
        else:
            form = TriageAssessmentForm(initial={
                'chief_complaint': visit.chief_complaint,
            })
    
    context = {
        'form': form,
        'visit': visit,
        'is_update': is_update,
        'page_title': f'Triage Assessment - {visit.patient.full_name}',
    }
    
    return render(request, 'visits/triage_assessment.html', context)


# =============================================================================
# ALL VISITS LIST
# =============================================================================

@login_required
def visits_list(request):
    """
    Display all patient visits with filtering
    """
    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    visit_type_filter = request.GET.get('visit_type', '')
    date_filter = request.GET.get('date', '')
    
    # Base queryset
    visits = PatientVisit.objects.select_related('patient').order_by('-visit_date', '-arrival_time')
    
    # Apply filters
    if search_query:
        visits = visits.filter(
            Q(visit_number__icontains=search_query) |
            Q(patient__patient_number__icontains=search_query) |
            Q(patient__first_name__icontains=search_query) |
            Q(patient__last_name__icontains=search_query) |
            Q(patient__phone_number__icontains=search_query)
        )
    
    if status_filter:
        visits = visits.filter(status=status_filter)
    
    if visit_type_filter:
        visits = visits.filter(visit_type=visit_type_filter)
    
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            visits = visits.filter(visit_date=filter_date)
        except ValueError:
            pass
    
    # Pagination
    paginator = Paginator(visits, 25)
    page_number = request.GET.get('page', 1)
    visits_page = paginator.get_page(page_number)
    
    # Statistics
    total_visits = visits.count()
    today_visits = PatientVisit.objects.filter(visit_date=timezone.now().date()).count()
    active_visits = visits.filter(status__in=['WAITING', 'TRIAGE', 'CONSULTATION']).count()
    
    context = {
        'visits': visits_page,
        'search_query': search_query,
        'status_filter': status_filter,
        'visit_type_filter': visit_type_filter,
        'date_filter': date_filter,
        'total_visits': total_visits,
        'today_visits': today_visits,
        'active_visits': active_visits,
        'page_title': 'All Visits',
    }
    
    return render(request, 'visits/visits_list.html', context)


# =============================================================================
# VISIT DETAIL VIEW
# =============================================================================

@login_required
def visits_detail(request, visit_number):
    """
    Display detailed visit information
    """
    visit = get_object_or_404(
        PatientVisit.objects.select_related('patient'),
        visit_number=visit_number
    )
    
    # Get related records
    try:
        triage = visit.triage
    except TriageAssessment.DoesNotExist:
        triage = None
    
    consultations = visit.consultations.select_related('doctor').order_by('-consultation_start')
    clinical_notes = visit.clinical_notes.select_related('clinician').order_by('-created_at')
    lab_orders = visit.lab_orders.select_related('test', 'ordered_by').order_by('-ordered_at')
    radiology_orders = visit.radiology_orders.select_related('test', 'ordered_by').order_by('-ordered_at')
    prescriptions = visit.prescriptions.prefetch_related('items__drug').order_by('-prescribed_at')
    
    try:
        invoice = visit.invoice
    except Invoice.DoesNotExist:
        invoice = None
    
    try:
        admission = visit.admission
    except:
        admission = None
    
    context = {
        'visit': visit,
        'triage': triage,
        'consultations': consultations,
        'clinical_notes': clinical_notes,
        'lab_orders': lab_orders,
        'radiology_orders': radiology_orders,
        'prescriptions': prescriptions,
        'invoice': invoice,
        'admission': admission,
        'page_title': f'Visit Details - {visit.visit_number}',
    }
    
    return render(request, 'visits/visits_detail.html', context)


# =============================================================================
# PATIENT QUEUE (Waiting for Consultation)
# =============================================================================

from django.db.models import Avg, ExpressionWrapper, F, DurationField
from django.db.models.functions import Extract

@login_required
def patient_queue(request):
    """
    Display patients waiting for consultation
    """
    # Get doctor's role
    user_role = request.user.role.name if request.user.role else None
    
    # Base query - patients ready for consultation
    queue = PatientVisit.objects.filter(
        status='CONSULTATION',
        visit_date=timezone.now().date()
    ).select_related('patient', 'triage').order_by('priority_level', 'arrival_time')
    
    # Get consultations in progress
    in_consultation = PatientVisit.objects.filter(
        status='CONSULTATION',
        visit_date=timezone.now().date(),
        consultations__consultation_end__isnull=True
    ).select_related('patient').distinct()
    
    # Statistics
    total_waiting = queue.count()
    
    # Calculate average wait time manually since wait_time_minutes is a property
    if total_waiting > 0:
        total_wait_minutes = sum(visit.wait_time_minutes for visit in queue)
        average_wait_time = round(total_wait_minutes / total_waiting, 1)
    else:
        average_wait_time = 0
    
    # Priority breakdown
    critical = queue.filter(priority_level=1).count()
    emergency = queue.filter(priority_level=2).count()
    urgent = queue.filter(priority_level=3).count()
    normal = queue.filter(priority_level__gte=4).count()
    
    context = {
        'queue': queue,
        'in_consultation': in_consultation,
        'total_waiting': total_waiting,
        'average_wait_time': average_wait_time,
        'critical': critical,
        'emergency': emergency,
        'urgent': urgent,
        'normal': normal,
        'page_title': 'Patient Queue',
    }
    
    return render(request, 'visits/patient_queue.html', context)


# =============================================================================
# UPDATE VISIT STATUS
# =============================================================================

@login_required
def visit_update_status(request, visit_number):
    """
    Update visit status (AJAX)
    """
    if request.method == 'POST':
        visit = get_object_or_404(PatientVisit, visit_number=visit_number)
        new_status = request.POST.get('status')
        
        if new_status in dict(PatientVisit.STATUS_CHOICES):
            old_status = visit.status
            visit.status = new_status
            
            # Set exit time if completing visit
            if new_status == 'COMPLETED' and not visit.exit_time:
                visit.exit_time = timezone.now()
            
            visit.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Visit status updated from {old_status} to {new_status}',
                'new_status': new_status,
                'new_status_display': visit.get_status_display()
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid status'
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


# =============================================================================
# PATIENT SEARCH API (for visit registration)
# =============================================================================

@login_required
def patient_search_for_visit(request):
    """
    API endpoint for patient search during visit registration
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    patients = Patient.objects.filter(
        Q(patient_number__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(id_number__icontains=query) |
        Q(phone_number__icontains=query),
        is_active=True
    )[:10]
    
    results = [{
        'id': p.id,
        'patient_number': p.patient_number,
        'name': p.full_name,
        'id_number': p.id_number or 'N/A',
        'phone': p.phone_number,
        'age': p.age,
        'gender': p.get_gender_display(),
        'last_visit': p.visits.order_by('-visit_date').first().visit_date.strftime('%Y-%m-%d') if p.visits.exists() else 'Never',
    } for p in patients]
    
    return JsonResponse({'results': results})


# =============================================================================
# DASHBOARD STATISTICS
# =============================================================================

@login_required
def visit_statistics(request):
    """
    Get visit statistics for dashboard
    """
    today = timezone.now().date()
    
    # Today's statistics
    today_visits = PatientVisit.objects.filter(visit_date=today)
    total_today = today_visits.count()
    emergency_today = today_visits.filter(visit_type='EMERGENCY').count()
    completed_today = today_visits.filter(status='COMPLETED').count()
    
    # This week's statistics
    week_start = today - timedelta(days=today.weekday())
    week_visits = PatientVisit.objects.filter(visit_date__gte=week_start)
    
    # Average wait time
    avg_wait = today_visits.aggregate(Avg('wait_time_minutes'))['wait_time_minutes__avg'] or 0
    
    stats = {
        'total_today': total_today,
        'emergency_today': emergency_today,
        'completed_today': completed_today,
        'week_total': week_visits.count(),
        'average_wait_time': round(avg_wait, 1),
        'pending_triage': today_visits.filter(status='WAITING').count(),
        'in_consultation': today_visits.filter(status='CONSULTATION').count(),
    }
    
    return JsonResponse(stats)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Sum, F
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from .models import Drug, DrugStock, DrugCategory
from django.utils import timezone


@login_required
def drug_inventory_list(request):
    """Main drug inventory view with search functionality"""
    search_query = request.GET.get('search', '')
    category_filter = request.GET.get('category', '')
    stock_status = request.GET.get('stock_status', '')
    
    # Base queryset with stock annotation
    drugs = Drug.objects.filter(is_active=True).select_related('category')
    
    # Apply search filter
    if search_query:
        drugs = drugs.filter(
            Q(name__icontains=search_query) |
            Q(generic_name__icontains=search_query) |
            Q(brand_name__icontains=search_query) |
            Q(drug_code__icontains=search_query)
        )
    
    # Apply category filter
    if category_filter:
        drugs = drugs.filter(category_id=category_filter)
    
    # Apply stock status filter
    if stock_status == 'low':
        drugs = [drug for drug in drugs if drug.needs_reorder]
    elif stock_status == 'out':
        drugs = [drug for drug in drugs if drug.current_stock == 0]
    elif stock_status == 'available':
        drugs = [drug for drug in drugs if drug.current_stock > drug.reorder_level]
    
    # Get all categories for filter dropdown
    categories = DrugCategory.objects.all().order_by('name')
    
    # For AJAX requests, return only the HTML fragment
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render(request, 'pharmacy/includes/drug_cards.html', {
            'drugs': drugs
        }).content.decode('utf-8')
        return JsonResponse({'html': html})
    
    context = {
        'drugs': drugs,
        'categories': categories,
        'search_query': search_query,
        'category_filter': category_filter,
        'stock_status': stock_status,
    }
    
    return render(request, 'pharmacy/drug_inventory.html', context)


@login_required
def drug_detail_ajax(request, drug_id):
    """AJAX endpoint to get drug details"""
    drug = get_object_or_404(Drug, id=drug_id)
    
    # Get stock batches
    stock_batches = DrugStock.objects.filter(
        drug=drug,
        quantity__gt=0
    ).order_by('expiry_date')
    
    # Get expiring soon batches (within 3 months)
    three_months_from_now = timezone.now().date() + timezone.timedelta(days=90)
    expiring_soon = stock_batches.filter(expiry_date__lte=three_months_from_now)
    
    data = {
        'id': drug.id,
        'name': drug.name,
        'generic_name': drug.generic_name,
        'brand_name': drug.brand_name,
        'drug_code': drug.drug_code,
        'category': drug.category.name if drug.category else 'No category',
        'form': drug.get_form_display(),
        'strength': drug.strength,
        'unit_price': str(drug.unit_price),
        'current_stock': drug.current_stock,
        'reorder_level': drug.reorder_level,
        'needs_reorder': drug.needs_reorder,
        'description': drug.description,
        'contraindications': drug.contraindications,
        'side_effects': drug.side_effects,
        'requires_prescription': drug.requires_prescription,
        'stock_batches': [
            {
                'batch_number': batch.batch_number,
                'quantity': batch.quantity,
                'expiry_date': batch.expiry_date.strftime('%Y-%m-%d'),
                'days_to_expiry': batch.days_to_expiry,
                'is_expired': batch.is_expired,
                'supplier': batch.supplier_name,
            }
            for batch in stock_batches
        ],
        'expiring_soon_count': expiring_soon.count(),
        'last_updated': drug.updated_at.strftime('%Y-%m-%d %H:%M'),
    }
    
    return JsonResponse(data)


@login_required
def drug_update_ajax(request, drug_id):
    """AJAX endpoint to update drug details"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)
    
    try:
        drug = get_object_or_404(Drug, id=drug_id)
        
        import json
        data = json.loads(request.body)
        
        # Update drug fields
        drug.name = data.get('name', drug.name)
        drug.generic_name = data.get('generic_name', drug.generic_name)
        drug.brand_name = data.get('brand_name', drug.brand_name)
        drug.strength = data.get('strength', drug.strength)
        drug.unit_price = data.get('unit_price', drug.unit_price)
        drug.reorder_level = data.get('reorder_level', drug.reorder_level)
        drug.description = data.get('description', drug.description)
        drug.contraindications = data.get('contraindications', drug.contraindications)
        drug.side_effects = data.get('side_effects', drug.side_effects)
        
        drug.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Drug updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
def drug_create(request):
    """Create new drug"""
    if request.method == 'POST':
        try:
            # Get category
            category_id = request.POST.get('category')
            category = get_object_or_404(DrugCategory, id=category_id) if category_id else None
            
            # Create drug
            drug = Drug.objects.create(
                name=request.POST.get('name'),
                generic_name=request.POST.get('generic_name'),
                brand_name=request.POST.get('brand_name', ''),
                drug_code=request.POST.get('drug_code'),
                category=category,
                form=request.POST.get('form'),
                strength=request.POST.get('strength'),
                unit_price=request.POST.get('unit_price'),
                reorder_level=request.POST.get('reorder_level', 10),
                description=request.POST.get('description', ''),
                contraindications=request.POST.get('contraindications', ''),
                side_effects=request.POST.get('side_effects', ''),
                requires_prescription=request.POST.get('requires_prescription') == 'on',
            )
            
            messages.success(request, f'Drug "{drug.name}" created successfully!')
            return redirect('drug-inventory')
            
        except Exception as e:
            messages.error(request, f'Error creating drug: {str(e)}')
    
    categories = DrugCategory.objects.all().order_by('name')
    context = {
        'categories': categories,
        'drug_forms': Drug.DRUG_FORM_CHOICES,
    }
    return render(request, 'pharmacy/drug_create.html', context)


@login_required
def drug_delete(request, drug_id):
    """Soft delete drug (set is_active to False)"""
    if request.method == 'POST':
        drug = get_object_or_404(Drug, id=drug_id)
        drug.is_active = False
        drug.save()
        messages.success(request, f'Drug "{drug.name}" deleted successfully!')
    
    return redirect('drug-inventory')


@login_required
def add_stock(request, drug_id):
    """Add new stock batch for a drug"""
    drug = get_object_or_404(Drug, id=drug_id)
    
    if request.method == 'POST':
        try:
            stock = DrugStock.objects.create(
                drug=drug,
                batch_number=request.POST.get('batch_number'),
                quantity=request.POST.get('quantity'),
                unit_cost=request.POST.get('unit_cost'),
                manufacture_date=request.POST.get('manufacture_date'),
                expiry_date=request.POST.get('expiry_date'),
                supplier_name=request.POST.get('supplier_name'),
                received_by=request.user,
                notes=request.POST.get('notes', ''),
            )
            
            messages.success(request, f'Stock added successfully! Batch: {stock.batch_number}')
            return redirect('drug-inventory')
            
        except Exception as e:
            messages.error(request, f'Error adding stock: {str(e)}')
    
    context = {
        'drug': drug,
    }
    return render(request, 'pharmacy/add_stock.html', context)


@login_required
def low_stock_report(request):
    """Report of drugs with low stock"""
    drugs = Drug.objects.filter(is_active=True).select_related('category')
    low_stock_drugs = [drug for drug in drugs if drug.needs_reorder]
    
    context = {
        'drugs': low_stock_drugs,
        'total_items': len(low_stock_drugs),
    }
    return render(request, 'pharmacy/low_stock_report.html', context)


@login_required
def expiring_stock_report(request):
    """Report of stock expiring soon"""
    days_ahead = int(request.GET.get('days', 90))
    cutoff_date = timezone.now().date() + timezone.timedelta(days=days_ahead)
    
    expiring_stock = DrugStock.objects.filter(
        quantity__gt=0,
        expiry_date__lte=cutoff_date,
        expiry_date__gte=timezone.now().date()
    ).select_related('drug', 'drug__category').order_by('expiry_date')
    
    context = {
        'expiring_stock': expiring_stock,
        'days_ahead': days_ahead,
        'cutoff_date': cutoff_date,
    }
    return render(request, 'pharmacy/expiring_stock_report.html', context)



# =============================================================================
# CLINICAL SERVICES VIEWS
# Views for Consultations and Clinical Notes
# =============================================================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count, Avg
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.http import require_http_methods
from decimal import Decimal

from .models import (
    PatientVisit, Consultation, ClinicalNote, Patient, 
    User, HospitalSettings, Invoice, InvoiceItem
)


# =============================================================================
# CONSULTATIONS
# =============================================================================

@login_required
def consultation_list(request):
    """
    List all consultations with filtering and search
    """
    # Base queryset
    consultations = Consultation.objects.select_related(
        'visit__patient', 'doctor'
    ).order_by('-consultation_start')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        consultations = consultations.filter(
            Q(visit__patient__first_name__icontains=search_query) |
            Q(visit__patient__last_name__icontains=search_query) |
            Q(visit__patient__patient_number__icontains=search_query) |
            Q(visit__visit_number__icontains=search_query) |
            Q(final_diagnosis__icontains=search_query)
        )
    
    # Filter by doctor
    doctor_filter = request.GET.get('doctor', '')
    if doctor_filter:
        consultations = consultations.filter(doctor_id=doctor_filter)
    
    # Filter by date
    date_filter = request.GET.get('date', '')
    if date_filter:
        consultations = consultations.filter(consultation_start__date=date_filter)
    
    # Filter by status (completed/ongoing)
    status_filter = request.GET.get('status', '')
    if status_filter == 'completed':
        consultations = consultations.filter(consultation_end__isnull=False)
    elif status_filter == 'ongoing':
        consultations = consultations.filter(consultation_end__isnull=True)
    
    # Pagination
    paginator = Paginator(consultations, 20)
    page_number = request.GET.get('page')
    consultations_page = paginator.get_page(page_number)
    
    # Get list of doctors for filter dropdown
    doctors = User.objects.filter(
        role__name__in=['DOCTOR', 'CLINICAL_OFFICER']
    ).order_by('first_name', 'last_name')
    
    # Statistics
    today = timezone.now().date()
    stats = {
        'total_today': Consultation.objects.filter(
            consultation_start__date=today
        ).count(),
        'completed_today': Consultation.objects.filter(
            consultation_start__date=today,
            consultation_end__isnull=False
        ).count(),
        'ongoing': Consultation.objects.filter(
            consultation_end__isnull=True
        ).count(),
        'total_this_month': Consultation.objects.filter(
            consultation_start__month=today.month,
            consultation_start__year=today.year
        ).count(),
    }
    
    context = {
        'consultations': consultations_page,
        'doctors': doctors,
        'search_query': search_query,
        'doctor_filter': doctor_filter,
        'date_filter': date_filter,
        'status_filter': status_filter,
        'stats': stats,
        'page_title': 'Consultations',
    }
    
    return render(request, 'clinical/consultation_list.html', context)


@login_required
def consultation_create(request):
    """
    Create a new consultation
    """
    if request.method == 'POST':
        try:
            # Get patient visit
            visit_id = request.POST.get('visit')
            visit = get_object_or_404(PatientVisit, id=visit_id)
            
            # Create consultation
            consultation = Consultation.objects.create(
                visit=visit,
                doctor=request.user,
                chief_complaint=request.POST.get('chief_complaint'),
                history_of_illness=request.POST.get('history_of_illness'),
                past_medical_history=request.POST.get('past_medical_history', ''),
                physical_examination=request.POST.get('physical_examination'),
                provisional_diagnosis=request.POST.get('provisional_diagnosis', ''),
                final_diagnosis=request.POST.get('final_diagnosis'),
                differential_diagnosis=request.POST.get('differential_diagnosis', ''),
                treatment_plan=request.POST.get('treatment_plan'),
                follow_up_instructions=request.POST.get('follow_up_instructions', ''),
                follow_up_date=request.POST.get('follow_up_date') or None,
                admission_required=request.POST.get('admission_required') == 'on',
                referral_required=request.POST.get('referral_required') == 'on',
                referral_facility=request.POST.get('referral_facility', ''),
            )
            
            # Update visit status
            visit.status = 'CONSULTATION'
            visit.save()
            
            messages.success(request, f'Consultation created successfully for {visit.patient.full_name}')
            return redirect('consultation-detail', consultation.id)
            
        except Exception as e:
            messages.error(request, f'Error creating consultation: {str(e)}')
            return redirect('consultation-list')
    
    # GET request - show form
    # Get visits waiting for consultation
    waiting_visits = PatientVisit.objects.filter(
        status__in=['WAITING', 'TRIAGE'],
        visit_date=timezone.now().date()
    ).select_related('patient', 'triage').order_by('priority_level', 'arrival_time')
    
    context = {
        'waiting_visits': waiting_visits,
        'page_title': 'New Consultation',
    }
    
    return render(request, 'clinical/consultation_create.html', context)


@login_required
def consultation_detail(request, consultation_id):
    """
    View consultation details
    """
    consultation = get_object_or_404(
        Consultation.objects.select_related(
            'visit__patient', 'doctor', 'visit__triage'
        ),
        id=consultation_id
    )
    
    # Get related data
    clinical_notes = consultation.visit.clinical_notes.all().order_by('-created_at')
    lab_orders = consultation.visit.lab_orders.select_related('test').order_by('-ordered_at')
    radiology_orders = consultation.visit.radiology_orders.select_related('test').order_by('-ordered_at')
    prescriptions = consultation.visit.prescriptions.prefetch_related('items__drug').order_by('-prescribed_at')
    
    context = {
        'consultation': consultation,
        'clinical_notes': clinical_notes,
        'lab_orders': lab_orders,
        'radiology_orders': radiology_orders,
        'prescriptions': prescriptions,
        'page_title': f'Consultation - {consultation.visit.visit_number}',
    }
    
    return render(request, 'clinical/consultation_detail.html', context)


@login_required
def consultation_update(request, consultation_id):
    """
    Update an existing consultation
    """
    consultation = get_object_or_404(Consultation, id=consultation_id)
    
    # Check permissions - only the consulting doctor or admin can update
    if request.user != consultation.doctor and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to update this consultation.')
        return redirect('consultation-detail', consultation_id)
    
    if request.method == 'POST':
        try:
            # Update consultation fields
            consultation.chief_complaint = request.POST.get('chief_complaint')
            consultation.history_of_illness = request.POST.get('history_of_illness')
            consultation.past_medical_history = request.POST.get('past_medical_history', '')
            consultation.physical_examination = request.POST.get('physical_examination')
            consultation.provisional_diagnosis = request.POST.get('provisional_diagnosis', '')
            consultation.final_diagnosis = request.POST.get('final_diagnosis')
            consultation.differential_diagnosis = request.POST.get('differential_diagnosis', '')
            consultation.treatment_plan = request.POST.get('treatment_plan')
            consultation.follow_up_instructions = request.POST.get('follow_up_instructions', '')
            consultation.follow_up_date = request.POST.get('follow_up_date') or None
            consultation.admission_required = request.POST.get('admission_required') == 'on'
            consultation.referral_required = request.POST.get('referral_required') == 'on'
            consultation.referral_facility = request.POST.get('referral_facility', '')
            
            # Mark as completed if specified
            if request.POST.get('mark_completed') == 'on':
                consultation.consultation_end = timezone.now()
            
            consultation.save()
            
            messages.success(request, 'Consultation updated successfully')
            return redirect('consultation-detail', consultation_id)
            
        except Exception as e:
            messages.error(request, f'Error updating consultation: {str(e)}')
    
    context = {
        'consultation': consultation,
        'page_title': f'Update Consultation - {consultation.visit.visit_number}',
    }
    
    return render(request, 'clinical/consultation_update.html', context)


@login_required
def consultation_complete(request, consultation_id):
    """
    Mark consultation as completed
    """
    consultation = get_object_or_404(Consultation, id=consultation_id)
    
    # Check permissions
    if request.user != consultation.doctor and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to complete this consultation.')
        return redirect('consultation-detail', consultation_id)
    
    if request.method == 'POST':
        consultation.consultation_end = timezone.now()
        consultation.save()
        
        # Update visit status
        consultation.visit.status = 'COMPLETED'
        consultation.visit.exit_time = timezone.now()
        consultation.visit.save()
        
        messages.success(request, 'Consultation marked as completed')
        return redirect('consultation-detail', consultation_id)
    
    return redirect('consultation-detail', consultation_id)


# =============================================================================
# CLINICAL NOTES
# =============================================================================

@login_required
def clinical_notes_list(request):
    """
    List all clinical notes with filtering
    """
    # Base queryset
    notes = ClinicalNote.objects.select_related(
        'visit__patient', 'clinician'
    ).order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        notes = notes.filter(
            Q(visit__patient__first_name__icontains=search_query) |
            Q(visit__patient__last_name__icontains=search_query) |
            Q(visit__patient__patient_number__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(content__icontains=search_query)
        )
    
    # Filter by note type
    note_type = request.GET.get('note_type', '')
    if note_type:
        notes = notes.filter(note_type=note_type)
    
    # Filter by clinician
    clinician_filter = request.GET.get('clinician', '')
    if clinician_filter:
        notes = notes.filter(clinician_id=clinician_filter)
    
    # Filter by date
    date_filter = request.GET.get('date', '')
    if date_filter:
        notes = notes.filter(created_at__date=date_filter)
    
    # Pagination
    paginator = Paginator(notes, 20)
    page_number = request.GET.get('page')
    notes_page = paginator.get_page(page_number)
    
    # Get clinicians for filter
    clinicians = User.objects.filter(
        role__name__in=['DOCTOR', 'CLINICAL_OFFICER', 'NURSE']
    ).order_by('first_name', 'last_name')
    
    # Note types for filter
    note_types = ClinicalNote.NOTE_TYPE_CHOICES
    
    # Statistics
    today = timezone.now().date()
    stats = {
        'total_today': ClinicalNote.objects.filter(created_at__date=today).count(),
        'consultation_notes': ClinicalNote.objects.filter(note_type='CONSULTATION').count(),
        'progress_notes': ClinicalNote.objects.filter(note_type='PROGRESS').count(),
        'procedure_notes': ClinicalNote.objects.filter(note_type='PROCEDURE').count(),
    }
    
    context = {
        'notes': notes_page,
        'clinicians': clinicians,
        'note_types': note_types,
        'search_query': search_query,
        'note_type': note_type,
        'clinician_filter': clinician_filter,
        'date_filter': date_filter,
        'stats': stats,
        'page_title': 'Clinical Notes',
    }
    
    return render(request, 'clinical/clinical_notes_list.html', context)


@login_required
def clinical_note_create(request):
    """
    Create a new clinical note
    """
    if request.method == 'POST':
        try:
            # Get visit
            visit_id = request.POST.get('visit')
            visit = get_object_or_404(PatientVisit, id=visit_id)
            
            # Create clinical note
            note = ClinicalNote.objects.create(
                visit=visit,
                note_type=request.POST.get('note_type'),
                clinician=request.user,
                subject=request.POST.get('subject'),
                content=request.POST.get('content'),
            )
            
            messages.success(request, 'Clinical note created successfully')
            return redirect('clinical-note-detail', note.id)
            
        except Exception as e:
            messages.error(request, f'Error creating clinical note: {str(e)}')
            return redirect('clinical-notes-list')
    
    # GET request - show form
    # Get active visits
    active_visits = PatientVisit.objects.filter(
        status__in=['CONSULTATION', 'ADMITTED'],
        visit_date__gte=timezone.now().date() - timezone.timedelta(days=7)
    ).select_related('patient').order_by('-visit_date')
    
    # Note types
    note_types = ClinicalNote.NOTE_TYPE_CHOICES
    
    context = {
        'active_visits': active_visits,
        'note_types': note_types,
        'page_title': 'New Clinical Note',
    }
    
    return render(request, 'clinical/clinical_note_create.html', context)


@login_required
def clinical_note_detail(request, note_id):
    """
    View clinical note details
    """
    note = get_object_or_404(
        ClinicalNote.objects.select_related(
            'visit__patient', 'clinician'
        ),
        id=note_id
    )
    
    # Get other notes for this visit
    related_notes = ClinicalNote.objects.filter(
        visit=note.visit
    ).exclude(id=note_id).order_by('-created_at')
    
    context = {
        'note': note,
        'related_notes': related_notes,
        'page_title': f'Clinical Note - {note.subject}',
    }
    
    return render(request, 'clinical/clinical_note_detail.html', context)


@login_required
def clinical_note_update(request, note_id):
    """
    Update clinical note
    """
    note = get_object_or_404(ClinicalNote, id=note_id)
    
    # Check permissions - only the author or admin can update
    if request.user != note.clinician and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to update this note.')
        return redirect('clinical-note-detail', note_id)
    
    if request.method == 'POST':
        try:
            note.note_type = request.POST.get('note_type')
            note.subject = request.POST.get('subject')
            note.content = request.POST.get('content')
            note.save()
            
            messages.success(request, 'Clinical note updated successfully')
            return redirect('clinical-note-detail', note_id)
            
        except Exception as e:
            messages.error(request, f'Error updating clinical note: {str(e)}')
    
    # Note types
    note_types = ClinicalNote.NOTE_TYPE_CHOICES
    
    context = {
        'note': note,
        'note_types': note_types,
        'page_title': f'Update Clinical Note - {note.subject}',
    }
    
    return render(request, 'clinical/clinical_note_update.html', context)


@login_required
def clinical_note_delete(request, note_id):
    """
    Delete clinical note
    """
    note = get_object_or_404(ClinicalNote, id=note_id)
    
    # Check permissions - only the author or admin can delete
    if request.user != note.clinician and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to delete this note.')
        return redirect('clinical-note-detail', note_id)
    
    if request.method == 'POST':
        visit_id = note.visit.id
        note.delete()
        
        messages.success(request, 'Clinical note deleted successfully')
        return redirect('clinical-notes-list')
    
    return redirect('clinical-note-detail', note_id)


# =============================================================================
# AJAX ENDPOINTS
# =============================================================================

@login_required
@require_http_methods(["GET"])
def patient_search_ajax(request):
    """
    AJAX endpoint for patient search
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    # Search patients
    patients = Patient.objects.filter(
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(patient_number__icontains=query) |
        Q(id_number__icontains=query) |
        Q(phone_number__icontains=query)
    ).order_by('-registration_date')[:10]
    
    results = []
    for patient in patients:
        results.append({
            'id': patient.id,
            'text': f"{patient.full_name} (ID: {patient.id_number or 'N/A'})",
            'patient_number': patient.patient_number,
            'name': patient.full_name,
            'id_number': patient.id_number or 'N/A',
            'dob': patient.date_of_birth.strftime('%Y-%m-%d'),
            'gender': patient.get_gender_display(),
            'phone': patient.phone_number,
        })
    
    return JsonResponse({'results': results})


@login_required
@require_http_methods(["GET"])
def visit_search_ajax(request):
    """
    AJAX endpoint for visit search
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    # Search visits
    visits = PatientVisit.objects.filter(
        Q(visit_number__icontains=query) |
        Q(patient__first_name__icontains=query) |
        Q(patient__last_name__icontains=query) |
        Q(patient__patient_number__icontains=query)
    ).select_related('patient').order_by('-visit_date')[:10]
    
    results = []
    for visit in visits:
        results.append({
            'id': visit.id,
            'text': f"{visit.visit_number} - {visit.patient.full_name}",
            'visit_number': visit.visit_number,
            'patient_name': visit.patient.full_name,
            'visit_date': visit.visit_date.strftime('%Y-%m-%d'),
            'status': visit.get_status_display(),
        })
    
    return JsonResponse({'results': results})



# receptionist/views.py
"""
Receptionist Portal Views
Handles patient registration, queue management, and appointments
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.utils import timezone
from django.http import JsonResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta

from .models import (
    Patient, PatientVisit, Appointment, User, 
    Department, HospitalSettings
)
from .forms import (
    PatientRegistrationForm, PatientVisitForm, 
    AppointmentForm, PatientSearchForm
)


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def receptionist_dashboard_view(request):
    """Receptionist dashboard with key metrics and quick actions"""
    
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    
    # Get today's statistics
    todays_appointments = Appointment.objects.filter(
        appointment_datetime__date=today,
        status__in=['SCHEDULED', 'CONFIRMED', 'ARRIVED']
    ).select_related('patient', 'doctor')
    
    todays_visits = PatientVisit.objects.filter(
        visit_date=today
    ).select_related('patient')
    
    # Get new patients this week
    new_patients_this_week = Patient.objects.filter(
        registration_date__gte=week_start
    ).count()
    
    # Recent patient registrations
    recent_patients = Patient.objects.filter(
        is_active=True
    ).order_by('-registration_date')[:6]
    
    # Available doctors (mock data - adjust based on your logic)
    available_doctors = User.objects.filter(
        role__name__in=['DOCTOR', 'CLINICAL_OFFICER'],
        is_active_staff=True
    )[:5]
    
    # Mock announcements (you can create an Announcement model)
    announcements = []
    
    # Pending prescriptions (mock - adjust based on your prescription model)
    pending_prescriptions = []
    
    context = {
        'todays_appointments': todays_appointments,
        'todays_visits': todays_visits,
        'new_patients_this_week': new_patients_this_week,
        'recent_patients': recent_patients,
        'available_doctors': available_doctors,
        'announcements': announcements,
        'pending_prescriptions': pending_prescriptions,
    }
    
    return render(request, 'receptionist/dashboard.html', context)


# =============================================================================
# PATIENT MANAGEMENT
# =============================================================================

@login_required
def receptionist_register_patient_view(request):
    """Register a new patient"""
    
    if request.method == 'POST':
        form = PatientRegistrationForm(request.POST)
        if form.is_valid():
            patient = form.save(commit=False)
            patient.registered_by = request.user
            patient.save()
            
            messages.success(
                request, 
                f'Patient {patient.full_name} registered successfully! '
                f'Patient Number: {patient.patient_number}'
            )
            
            # Redirect to create visit or back to form
            if 'save_and_visit' in request.POST:
                return redirect('receptionist_create_visit', patient_id=patient.id)
            else:
                return redirect('receptionist_register_patient')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PatientRegistrationForm()
    
    context = {
        'form': form,
        'title': 'Register New Patient'
    }
    
    return render(request, 'receptionist/register_patient.html', context)


@login_required
def receptionist_patient_records_view(request):
    """View all patient records with search and filter"""
    
    # Get search parameters
    search_query = request.GET.get('search', '')
    filter_status = request.GET.get('status', 'all')
    
    # Base queryset
    patients = Patient.objects.all().order_by('-registration_date')
    
    # Apply filters
    if search_query:
        patients = patients.filter(
            Q(patient_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id_number__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    if filter_status == 'active':
        patients = patients.filter(is_active=True)
    elif filter_status == 'inactive':
        patients = patients.filter(is_active=False)
    
    # Pagination
    paginator = Paginator(patients, 20)  # 20 patients per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'filter_status': filter_status,
        'total_patients': patients.count(),
        'title': 'Patient Records'
    }
    
    return render(request, 'receptionist/patient_records.html', context)


@login_required
def receptionist_search_patient_view(request):
    """Advanced patient search"""
    
    patients = []
    search_performed = False
    
    if request.method == 'GET' and request.GET:
        form = PatientSearchForm(request.GET)
        if form.is_valid():
            search_performed = True
            patients = Patient.objects.all()
            
            # Apply filters
            if form.cleaned_data.get('patient_number'):
                patients = patients.filter(
                    patient_number__icontains=form.cleaned_data['patient_number']
                )
            
            if form.cleaned_data.get('id_number'):
                patients = patients.filter(
                    id_number=form.cleaned_data['id_number']
                )
            
            if form.cleaned_data.get('phone_number'):
                patients = patients.filter(
                    phone_number__icontains=form.cleaned_data['phone_number']
                )
            
            if form.cleaned_data.get('first_name'):
                patients = patients.filter(
                    first_name__icontains=form.cleaned_data['first_name']
                )
            
            if form.cleaned_data.get('last_name'):
                patients = patients.filter(
                    last_name__icontains=form.cleaned_data['last_name']
                )
            
            if form.cleaned_data.get('date_of_birth'):
                patients = patients.filter(
                    date_of_birth=form.cleaned_data['date_of_birth']
                )
            
            patients = patients.order_by('-registration_date')[:50]
    else:
        form = PatientSearchForm()
    
    context = {
        'form': form,
        'patients': patients,
        'search_performed': search_performed,
        'title': 'Search Patient'
    }
    
    return render(request, 'receptionist/search_patient.html', context)


@login_required
def receptionist_patient_detail_view(request, patient_number):
    """
    View detailed patient information using patient number
    """
    patient = get_object_or_404(
        Patient.objects.select_related('registered_by'),
        patient_number=patient_number,
        is_active=True
    )
    
    # Get patient's recent visits
    recent_visits = patient.visits.all().order_by('-visit_date')[:10]
    
    # Get patient's prescriptions
    recent_prescriptions = patient.visits.filter(
        prescriptions__isnull=False
    ).order_by('-visit_date')[:5]
    
    # Get patient's invoices
    recent_invoices = patient.invoices.all().order_by('-invoice_date')[:10]
    
    # Get patient's admissions
    admissions = patient.admissions.all().order_by('-admission_datetime')[:5]
    
    # Calculate statistics
    total_visits = patient.visits.count()
    total_amount_billed = sum(invoice.total_amount for invoice in patient.invoices.all())
    total_amount_paid = sum(invoice.amount_paid for invoice in patient.invoices.all())
    outstanding_balance = sum(invoice.balance for invoice in patient.invoices.all())
    
    context = {
        'patient': patient,
        'recent_visits': recent_visits,
        'recent_prescriptions': recent_prescriptions,
        'recent_invoices': recent_invoices,
        'admissions': admissions,
        'total_visits': total_visits,
        'total_amount_billed': total_amount_billed,
        'total_amount_paid': total_amount_paid,
        'outstanding_balance': outstanding_balance,
        'page_title': f'Patient: {patient.full_name}',
    }
    
    return render(request, 'receptionist/patient_detail.html', context)


@login_required
def receptionist_edit_patient_view(request, patient_number):
    """Edit patient information"""
    
    """
    Update patient information
    """
    patient = get_object_or_404(Patient, patient_number=patient_number, is_active=True)
    
    if request.method == 'POST':
        form = PatientRegistrationForm(request.POST, instance=patient)
        
        if form.is_valid():
            patient = form.save()
            messages.success(request, f'Patient {patient.full_name} updated successfully!')
            return redirect('patients-detail', patient_number=patient.patient_number)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PatientRegistrationForm(instance=patient)
    
    context = {
        'form': form,
        'patient': patient,
        'page_title': f'Edit Patient: {patient.full_name}',
    }
    
    return render(request, 'receptionist/edit_patient.html', context)


# =============================================================================
# QUEUE MANAGEMENT
# =============================================================================
# receptionist/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.db import transaction
from django.db.models import Q, Count, Avg, F, ExpressionWrapper, fields
from django.utils import timezone
from datetime import datetime, timedelta
import json

from .models import (
    Patient, PatientVisit, TriageAssessment, 
    User, Department, Notification
)


# =============================================================================
# QUEUE MANAGEMENT - MAIN VIEW
# =============================================================================

@login_required
def queue_management(request):
    """Main queue management dashboard"""
    today = timezone.now().date()
    
    # Get all active visits for today
    active_visits = PatientVisit.objects.filter(
        visit_date=today,
        status__in=['WAITING', 'TRIAGE', 'CONSULTATION', 'LABORATORY', 
                    'RADIOLOGY', 'PHARMACY', 'BILLING']
    ).select_related(
        'patient', 'triage'
    ).order_by('priority_level', 'queue_number')
    
    # Statistics
    stats = {
        'total_patients': active_visits.count(),
        'waiting': active_visits.filter(status='WAITING').count(),
        'in_triage': active_visits.filter(status='TRIAGE').count(),
        'in_consultation': active_visits.filter(status='CONSULTATION').count(),
        'average_wait_time': _calculate_average_wait_time(active_visits),
    }
    
    # Get departments for filtering
    departments = Department.objects.filter(
        is_active=True,
        name__in=['TRIAGE', 'OUTPATIENT', 'LABORATORY', 'RADIOLOGY', 'PHARMACY', 'BILLING']
    )
    
    # Get available doctors for assignment
    doctors = User.objects.filter(
        role__name__in=['DOCTOR', 'CLINICAL_OFFICER'],
        is_active_staff=True
    ).select_related('role', 'department')
    
    context = {
        'stats': stats,
        'departments': departments,
        'doctors': doctors,
        'page_title': 'Queue Management',
    }
    
    return render(request, 'receptionist/queue_management.html', context)


def _calculate_average_wait_time(visits):
    """Calculate average waiting time in minutes"""
    if not visits.exists():
        return 0
    
    total_minutes = 0
    count = 0
    
    for visit in visits:
        if visit.status != 'COMPLETED':
            minutes = visit.wait_time_minutes
            total_minutes += minutes
            count += 1
    
    return round(total_minutes / count) if count > 0 else 0


# =============================================================================
# AJAX API - GET QUEUE DATA
# =============================================================================

@login_required
@require_http_methods(["GET"])
def api_get_queue(request):
    """Get current queue data with filters"""
    today = timezone.now().date()
    
    # Get filters from request
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    search_query = request.GET.get('search', '')
    
    # Base queryset
    visits = PatientVisit.objects.filter(
        visit_date=today,
        status__in=['WAITING', 'TRIAGE', 'CONSULTATION', 'LABORATORY', 
                    'RADIOLOGY', 'PHARMACY', 'BILLING']
    ).select_related('patient', 'triage')
    
    # Apply filters
    if status_filter:
        visits = visits.filter(status=status_filter)
    
    if priority_filter:
        visits = visits.filter(priority_level=int(priority_filter))
    
    if search_query:
        visits = visits.filter(
            Q(patient__patient_number__icontains=search_query) |
            Q(patient__first_name__icontains=search_query) |
            Q(patient__last_name__icontains=search_query) |
            Q(patient__phone_number__icontains=search_query)
        )
    
    # Order by priority and queue number
    visits = visits.order_by('priority_level', 'queue_number')
    
    # Serialize data
    queue_data = []
    for visit in visits:
        triage = getattr(visit, 'triage', None)
        
        queue_data.append({
            'id': visit.id,
            'visit_number': visit.visit_number,
            'queue_number': visit.queue_number,
            'patient': {
                'id': visit.patient.id,
                'patient_number': visit.patient.patient_number,
                'full_name': visit.patient.full_name,
                'age': visit.patient.age,
                'gender': visit.patient.get_gender_display(),
                'phone': visit.patient.phone_number,
            },
            'visit_type': visit.get_visit_type_display(),
            'priority_level': visit.priority_level,
            'priority_display': _get_priority_display(visit.priority_level),
            'status': visit.status,
            'status_display': visit.get_status_display(),
            'chief_complaint': visit.chief_complaint,
            'arrival_time': visit.arrival_time.strftime('%H:%M'),
            'wait_time_minutes': visit.wait_time_minutes,
            'triage': {
                'done': triage is not None,
                'emergency_level': triage.emergency_level if triage else None,
                'temperature': str(triage.temperature) if triage else None,
                'bp': f"{triage.systolic_bp}/{triage.diastolic_bp}" if triage else None,
            } if triage else None,
        })
    
    return JsonResponse({
        'success': True,
        'data': queue_data,
        'count': len(queue_data),
        'timestamp': timezone.now().isoformat(),
    })


def _get_priority_display(level):
    """Get priority level display text and color"""
    priority_map = {
        1: {'text': 'Critical', 'color': 'danger', 'icon': 'exclamation-triangle-fill'},
        2: {'text': 'Emergency', 'color': 'danger', 'icon': 'exclamation-circle-fill'},
        3: {'text': 'Urgent', 'color': 'warning', 'icon': 'exclamation-circle'},
        4: {'text': 'Normal', 'color': 'info', 'icon': 'info-circle'},
        5: {'text': 'Low', 'color': 'secondary', 'icon': 'circle'},
    }
    return priority_map.get(level, priority_map[4])


# =============================================================================
# AJAX API - UPDATE VISIT STATUS
# =============================================================================

@login_required
@require_POST
def api_update_visit_status(request):
    """Update patient visit status"""
    try:
        data = json.loads(request.body)
        visit_id = data.get('visit_id')
        new_status = data.get('status')
        notes = data.get('notes', '')
        
        if not visit_id or not new_status:
            return JsonResponse({
                'success': False,
                'error': 'Missing required fields'
            }, status=400)
        
        visit = get_object_or_404(PatientVisit, id=visit_id)
        
        # Update status
        old_status = visit.status
        visit.status = new_status
        
        # Add notes if status is being completed or cancelled
        if new_status in ['COMPLETED', 'CANCELLED']:
            visit.exit_time = timezone.now()
            visit.exit_notes = notes
        
        visit.save()
        
        # Create notification for relevant staff
        _send_status_change_notification(visit, old_status, new_status)
        
        return JsonResponse({
            'success': True,
            'message': f'Status updated to {visit.get_status_display()}',
            'data': {
                'visit_id': visit.id,
                'status': visit.status,
                'status_display': visit.get_status_display(),
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def _send_status_change_notification(visit, old_status, new_status):
    """Send notification when visit status changes"""
    # Notify relevant departments
    notification_map = {
        'TRIAGE': 'NURSE',
        'CONSULTATION': 'DOCTOR',
        'LABORATORY': 'LAB_TECHNICIAN',
        'RADIOLOGY': 'RADIOLOGIST',
        'PHARMACY': 'PHARMACIST',
        'BILLING': 'CASHIER',
    }
    
    if new_status in notification_map:
        role_name = notification_map[new_status]
        staff_members = User.objects.filter(
            role__name=role_name,
            is_active_staff=True
        )
        
        for staff in staff_members:
            Notification.objects.create(
                recipient=staff,
                notification_type='GENERAL',
                title='New Patient in Queue',
                message=f'Patient {visit.patient.full_name} ({visit.visit_number}) is now in {visit.get_status_display()}',
                link_url=f'/queue/{visit.id}/'
            )


# =============================================================================
# AJAX API - UPDATE PRIORITY
# =============================================================================

@login_required
@require_POST
def api_update_priority(request):
    """Update patient visit priority level"""
    try:
        data = json.loads(request.body)
        visit_id = data.get('visit_id')
        priority_level = data.get('priority_level')
        reason = data.get('reason', '')
        
        if not visit_id or priority_level is None:
            return JsonResponse({
                'success': False,
                'error': 'Missing required fields'
            }, status=400)
        
        visit = get_object_or_404(PatientVisit, id=visit_id)
        
        old_priority = visit.priority_level
        visit.priority_level = int(priority_level)
        visit.save()
        
        # Log the change
        if reason:
            visit.chief_complaint += f"\n[Priority changed: {reason}]"
            visit.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Priority updated successfully',
            'data': {
                'visit_id': visit.id,
                'priority_level': visit.priority_level,
                'priority_display': _get_priority_display(visit.priority_level),
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# AJAX API - ASSIGN TO DOCTOR
# =============================================================================

@login_required
@require_POST
def api_assign_doctor(request):
    """Assign patient to a specific doctor"""
    try:
        data = json.loads(request.body)
        visit_id = data.get('visit_id')
        doctor_id = data.get('doctor_id')
        
        if not visit_id or not doctor_id:
            return JsonResponse({
                'success': False,
                'error': 'Missing required fields'
            }, status=400)
        
        visit = get_object_or_404(PatientVisit, id=visit_id)
        doctor = get_object_or_404(User, id=doctor_id)
        
        # Update visit status
        visit.status = 'CONSULTATION'
        visit.save()
        
        # Send notification to doctor
        Notification.objects.create(
            recipient=doctor,
            notification_type='GENERAL',
            title='New Patient Assigned',
            message=f'Patient {visit.patient.full_name} ({visit.visit_number}) has been assigned to you',
            link_url=f'/consultations/{visit.id}/'
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Patient assigned to Dr. {doctor.last_name}',
            'data': {
                'visit_id': visit.id,
                'doctor_name': doctor.get_full_name(),
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# AJAX API - CALL NEXT PATIENT
# =============================================================================

@login_required
@require_POST
def api_call_next_patient(request):
    """Call the next patient in queue"""
    try:
        data = json.loads(request.body)
        department = data.get('department', 'TRIAGE')
        
        today = timezone.now().date()
        
        # Get next waiting patient
        next_patient = PatientVisit.objects.filter(
            visit_date=today,
            status='WAITING'
        ).order_by('priority_level', 'queue_number').first()
        
        if not next_patient:
            return JsonResponse({
                'success': False,
                'message': 'No patients in waiting queue'
            })
        
        # Update status based on department
        status_map = {
            'TRIAGE': 'TRIAGE',
            'CONSULTATION': 'CONSULTATION',
            'LABORATORY': 'LABORATORY',
            'RADIOLOGY': 'RADIOLOGY',
            'PHARMACY': 'PHARMACY',
            'BILLING': 'BILLING',
        }
        
        next_patient.status = status_map.get(department, 'TRIAGE')
        next_patient.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Called patient: {next_patient.patient.full_name}',
            'data': {
                'visit_id': next_patient.id,
                'visit_number': next_patient.visit_number,
                'queue_number': next_patient.queue_number,
                'patient_name': next_patient.patient.full_name,
                'patient_number': next_patient.patient.patient_number,
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# AJAX API - GET VISIT DETAILS
# =============================================================================

@login_required
@require_http_methods(["GET"])
def api_get_visit_details(request, visit_id):
    """Get detailed information about a specific visit"""
    try:
        visit = get_object_or_404(
            PatientVisit.objects.select_related('patient', 'triage'),
            id=visit_id
        )
        
        patient = visit.patient
        triage = getattr(visit, 'triage', None)
        
        # Get recent consultations
        recent_consultations = visit.consultations.select_related('doctor').order_by('-consultation_start')[:5]
        
        # Get pending orders
        lab_orders = visit.lab_orders.filter(status__in=['ORDERED', 'RECEIVED', 'IN_PROGRESS']).count()
        rad_orders = visit.radiology_orders.filter(status__in=['ORDERED', 'SCHEDULED', 'IN_PROGRESS']).count()
        prescriptions = visit.prescriptions.filter(status__in=['PENDING', 'PARTIALLY_DISPENSED']).count()
        
        data = {
            'visit': {
                'id': visit.id,
                'visit_number': visit.visit_number,
                'visit_type': visit.get_visit_type_display(),
                'queue_number': visit.queue_number,
                'priority_level': visit.priority_level,
                'status': visit.status,
                'status_display': visit.get_status_display(),
                'chief_complaint': visit.chief_complaint,
                'arrival_time': visit.arrival_time.strftime('%Y-%m-%d %H:%M'),
                'wait_time_minutes': visit.wait_time_minutes,
            },
            'patient': {
                'id': patient.id,
                'patient_number': patient.patient_number,
                'full_name': patient.full_name,
                'age': patient.age,
                'gender': patient.get_gender_display(),
                'phone': patient.phone_number,
                'blood_group': patient.blood_group,
                'allergies': patient.allergies,
                'chronic_conditions': patient.chronic_conditions,
                'nhif_status': patient.get_nhif_status_display(),
                'nhif_number': patient.nhif_number,
            },
            'triage': {
                'done': triage is not None,
                'emergency_level': triage.get_emergency_level_display() if triage else None,
                'temperature': str(triage.temperature) if triage else None,
                'pulse': triage.pulse if triage else None,
                'bp': f"{triage.systolic_bp}/{triage.diastolic_bp}" if triage else None,
                'respiratory_rate': triage.respiratory_rate if triage else None,
                'oxygen_saturation': triage.oxygen_saturation if triage else None,
                'pain_scale': triage.pain_scale if triage else None,
            } if triage else None,
            'consultations': [
                {
                    'id': cons.id,
                    'doctor': cons.doctor.get_full_name(),
                    'date': cons.consultation_start.strftime('%Y-%m-%d %H:%M'),
                    'diagnosis': cons.final_diagnosis[:100] + '...' if len(cons.final_diagnosis) > 100 else cons.final_diagnosis,
                }
                for cons in recent_consultations
            ],
            'pending_orders': {
                'laboratory': lab_orders,
                'radiology': rad_orders,
                'prescriptions': prescriptions,
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# AJAX API - GET QUEUE STATISTICS
# =============================================================================

@login_required
@require_http_methods(["GET"])
def api_get_queue_stats(request):
    """Get real-time queue statistics"""
    try:
        today = timezone.now().date()
        
        # Get all visits for today
        all_visits = PatientVisit.objects.filter(visit_date=today)
        active_visits = all_visits.filter(
            status__in=['WAITING', 'TRIAGE', 'CONSULTATION', 'LABORATORY', 
                        'RADIOLOGY', 'PHARMACY', 'BILLING']
        )
        
        # Calculate statistics
        stats = {
            'total_today': all_visits.count(),
            'active_patients': active_visits.count(),
            'completed': all_visits.filter(status='COMPLETED').count(),
            'by_status': {
                'waiting': active_visits.filter(status='WAITING').count(),
                'triage': active_visits.filter(status='TRIAGE').count(),
                'consultation': active_visits.filter(status='CONSULTATION').count(),
                'laboratory': active_visits.filter(status='LABORATORY').count(),
                'radiology': active_visits.filter(status='RADIOLOGY').count(),
                'pharmacy': active_visits.filter(status='PHARMACY').count(),
                'billing': active_visits.filter(status='BILLING').count(),
            },
            'by_priority': {
                'critical': active_visits.filter(priority_level=1).count(),
                'emergency': active_visits.filter(priority_level=2).count(),
                'urgent': active_visits.filter(priority_level=3).count(),
                'normal': active_visits.filter(priority_level=4).count(),
                'low': active_visits.filter(priority_level=5).count(),
            },
            'by_visit_type': {
                'outpatient': all_visits.filter(visit_type='OUTPATIENT').count(),
                'emergency': all_visits.filter(visit_type='EMERGENCY').count(),
                'referral': all_visits.filter(visit_type='REFERRAL').count(),
            },
            'average_wait_time': _calculate_average_wait_time(active_visits),
        }
        
        return JsonResponse({
            'success': True,
            'data': stats,
            'timestamp': timezone.now().isoformat(),
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# AJAX API - SEARCH PATIENTS IN QUEUE
# =============================================================================

@login_required
@require_http_methods(["GET"])
def api_search_queue(request):
    """Search for patients in today's queue"""
    try:
        query = request.GET.get('q', '').strip()
        
        if len(query) < 2:
            return JsonResponse({
                'success': False,
                'error': 'Search query must be at least 2 characters'
            }, status=400)
        
        today = timezone.now().date()
        
        visits = PatientVisit.objects.filter(
            visit_date=today,
            status__in=['WAITING', 'TRIAGE', 'CONSULTATION', 'LABORATORY', 
                        'RADIOLOGY', 'PHARMACY', 'BILLING']
        ).filter(
            Q(visit_number__icontains=query) |
            Q(patient__patient_number__icontains=query) |
            Q(patient__first_name__icontains=query) |
            Q(patient__last_name__icontains=query) |
            Q(patient__phone_number__icontains=query) |
            Q(patient__id_number__icontains=query)
        ).select_related('patient', 'triage')[:20]
        
        results = [
            {
                'visit_id': visit.id,
                'visit_number': visit.visit_number,
                'queue_number': visit.queue_number,
                'patient_number': visit.patient.patient_number,
                'patient_name': visit.patient.full_name,
                'status': visit.get_status_display(),
                'priority': _get_priority_display(visit.priority_level),
            }
            for visit in visits
        ]
        
        return JsonResponse({
            'success': True,
            'data': results,
            'count': len(results)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)



@login_required
def receptionist_create_visit_view(request, patient_id=None):
    """Create a new patient visit"""
    
    patient = None
    if patient_id:
        patient = get_object_or_404(Patient, id=patient_id)
    
    if request.method == 'POST':
        form = PatientVisitForm(request.POST)
        if form.is_valid():
            visit = form.save(commit=False)
            
            # If patient_id provided, use it
            if patient:
                visit.patient = patient
            
            visit.save()
            
            messages.success(
                request,
                f'Visit created successfully! Queue Number: {visit.queue_number}'
            )
            
            return redirect('receptionist_queue_management')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        initial_data = {}
        if patient:
            initial_data['patient'] = patient
        form = PatientVisitForm(initial=initial_data)
    
    context = {
        'form': form,
        'patient': patient,
        'title': 'Create Patient Visit'
    }
    
    return render(request, 'receptionist/create_visit.html', context)


# =============================================================================
# APPOINTMENTS
# =============================================================================

@login_required
def receptionist_schedule_appointment_view(request):
    """Schedule a new appointment"""
    
    if request.method == 'POST':
        form = AppointmentForm(request.POST)
        if form.is_valid():
            appointment = form.save()
            
            messages.success(
                request,
                f'Appointment scheduled for {appointment.patient.full_name} '
                f'on {appointment.appointment_datetime.strftime("%B %d, %Y at %I:%M %p")}'
            )
            
            # TODO: Send SMS reminder to patient
            
            return redirect('receptionist_view_appointments')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AppointmentForm()
    
    # Get available doctors
    doctors = User.objects.filter(
        role__name__in=['DOCTOR', 'CLINICAL_OFFICER'],
        is_active_staff=True
    )
    
    context = {
        'form': form,
        'doctors': doctors,
        'title': 'Schedule Appointment'
    }
    
    return render(request, 'receptionist/schedule_appointment.html', context)


@login_required
def receptionist_view_appointments_view(request):
    """View all appointments with filters"""
    
    # Get filter parameters
    date_filter = request.GET.get('date', 'today')
    status_filter = request.GET.get('status', 'all')
    doctor_id = request.GET.get('doctor', '')
    
    # Base queryset
    appointments = Appointment.objects.all().select_related('patient', 'doctor')
    
    # Apply date filter
    today = timezone.now().date()
    if date_filter == 'today':
        appointments = appointments.filter(appointment_datetime__date=today)
    elif date_filter == 'tomorrow':
        tomorrow = today + timedelta(days=1)
        appointments = appointments.filter(appointment_datetime__date=tomorrow)
    elif date_filter == 'week':
        week_end = today + timedelta(days=7)
        appointments = appointments.filter(
            appointment_datetime__date__gte=today,
            appointment_datetime__date__lte=week_end
        )
    elif date_filter == 'month':
        month_end = today + timedelta(days=30)
        appointments = appointments.filter(
            appointment_datetime__date__gte=today,
            appointment_datetime__date__lte=month_end
        )
    
    # Apply status filter
    if status_filter != 'all':
        appointments = appointments.filter(status=status_filter)
    
    # Apply doctor filter
    if doctor_id:
        appointments = appointments.filter(doctor_id=doctor_id)
    
    appointments = appointments.order_by('appointment_datetime')
    
    # Pagination
    paginator = Paginator(appointments, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get doctors for filter dropdown
    doctors = User.objects.filter(
        role__name__in=['DOCTOR', 'CLINICAL_OFFICER'],
        is_active_staff=True
    )
    
    context = {
        'page_obj': page_obj,
        'date_filter': date_filter,
        'status_filter': status_filter,
        'doctor_id': doctor_id,
        'doctors': doctors,
        'total_appointments': appointments.count(),
        'title': 'View Appointments'
    }
    
    return render(request, 'receptionist/view_appointments.html', context)


@login_required
def receptionist_appointment_detail_view(request, appointment_id):
    """View appointment details"""
    
    appointment = get_object_or_404(
        Appointment.objects.select_related('patient', 'doctor'),
        id=appointment_id
    )
    
    context = {
        'appointment': appointment,
        'title': f'Appointment: {appointment.appointment_number}'
    }
    
    return render(request, 'receptionist/appointment_detail.html', context)


@login_required
def receptionist_update_appointment_status_view(request, appointment_id):
    """Update appointment status (AJAX)"""
    
    if request.method == 'POST':
        appointment = get_object_or_404(Appointment, id=appointment_id)
        new_status = request.POST.get('status')
        
        if new_status in dict(Appointment.STATUS_CHOICES).keys():
            appointment.status = new_status
            appointment.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Appointment status updated to {appointment.get_status_display()}'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid status'
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


@login_required
def receptionist_cancel_appointment_view(request, appointment_id):
    """Cancel an appointment"""
    
    appointment = get_object_or_404(Appointment, id=appointment_id)
    
    if request.method == 'POST':
        reason = request.POST.get('cancellation_reason', '')
        appointment.status = 'CANCELLED'
        appointment.notes = f"Cancelled: {reason}\n{appointment.notes}"
        appointment.save()
        
        messages.success(request, f'Appointment {appointment.appointment_number} cancelled.')
        
        # TODO: Send SMS to patient about cancellation
        
        return redirect('receptionist_view_appointments')
    
    context = {
        'appointment': appointment,
        'title': 'Cancel Appointment'
    }
    
    return render(request, 'receptionist/cancel_appointment.html', context)


# =============================================================================
# REPORTS & STATISTICS
# =============================================================================

@login_required
def receptionist_daily_report_view(request):
    """View daily statistics and reports"""
    
    today = timezone.now().date()
    
    # Daily statistics
    stats = {
        'total_visits': PatientVisit.objects.filter(visit_date=today).count(),
        'new_patients': Patient.objects.filter(registration_date__date=today).count(),
        'completed_visits': PatientVisit.objects.filter(
            visit_date=today, 
            status='COMPLETED'
        ).count(),
        'appointments_today': Appointment.objects.filter(
            appointment_datetime__date=today
        ).count(),
        'emergency_visits': PatientVisit.objects.filter(
            visit_date=today,
            visit_type='EMERGENCY'
        ).count(),
    }
    
    # Visit type breakdown
    visit_types = PatientVisit.objects.filter(
        visit_date=today
    ).values('visit_type').annotate(count=Count('id'))
    
    # Hourly distribution
    hourly_visits = PatientVisit.objects.filter(
        visit_date=today
    ).extra(select={'hour': 'EXTRACT(hour FROM arrival_time)'}).values('hour').annotate(count=Count('id')).order_by('hour')
    
    context = {
        'stats': stats,
        'visit_types': visit_types,
        'hourly_visits': hourly_visits,
        'report_date': today,
        'title': 'Daily Report'
    }
    
    return render(request, 'receptionist/daily_report.html', context)


# =============================================================================
# UTILITY VIEWS
# =============================================================================

@login_required
def receptionist_print_queue_ticket_view(request, visit_id):
    """Print queue ticket for patient"""
    
    visit = get_object_or_404(PatientVisit, id=visit_id)
    hospital_settings = HospitalSettings.load()
    
    context = {
        'visit': visit,
        'hospital': hospital_settings,
        'print_time': timezone.now()
    }
    
    return render(request, 'receptionist/print_queue_ticket.html', context)


@login_required
def receptionist_check_in_patient_view(request, appointment_id):
    """Check in a patient with appointment"""
    
    appointment = get_object_or_404(Appointment, id=appointment_id)
    
    if request.method == 'POST':
        # Create visit from appointment
        visit = PatientVisit.objects.create(
            patient=appointment.patient,
            visit_type='OUTPATIENT',
            chief_complaint=appointment.reason,
            status='WAITING'
        )
        
        # Update appointment status
        appointment.status = 'ARRIVED'
        appointment.save()
        
        messages.success(
            request,
            f'{appointment.patient.full_name} checked in. Queue Number: {visit.queue_number}'
        )
        
        return redirect('receptionist_queue_management')
    
    context = {
        'appointment': appointment,
        'title': 'Check In Patient'
    }
    
    return render(request, 'receptionist/check_in_patient.html', context)